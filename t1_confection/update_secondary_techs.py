"""
Update Secondary Techs from Editor Template

This script reads the Secondary_Techs_Editor.xlsx file and applies
the changes to the corresponding A-O_Parametrization.xlsx files.

Usage:
    python t1_confection/update_secondary_techs.py
"""
import openpyxl
import sys
from pathlib import Path
from datetime import datetime
import shutil


class SecondaryTechsUpdater:
    def __init__(self, editor_path, base_path):
        self.editor_path = editor_path
        self.base_path = base_path
        self.scenarios = ["BAU", "NDC", "NDC+ELC", "NDC_NoRPO"]
        self.log_lines = []
        self.changes_applied = 0
        self.rows_failed = 0

    def log(self, message, level="INFO"):
        """Add message to log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{timestamp}] {level}: {message}"
        self.log_lines.append(log_line)
        print(log_line)

    def read_editor_file(self):
        """
        Read and parse the editor Excel file

        Returns:
            list of dicts with editing instructions
        """
        self.log("Reading editor file...")

        if not self.editor_path.exists():
            raise FileNotFoundError(f"Editor file not found: {self.editor_path}")

        wb = openpyxl.load_workbook(self.editor_path, data_only=True)

        if 'Editor' not in wb.sheetnames:
            raise ValueError("'Editor' sheet not found in template file")

        ws = wb['Editor']

        # Read header to get year columns
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                break

        # Identify year columns (after Scenario, Country, Tech.Name, Tech, Parameter)
        # Years start from column 6 (F)
        year_columns = []
        for idx, header in enumerate(headers[5:], 6):  # Start from column 6 (F)
            if header.isdigit():
                year_columns.append((idx, int(header)))

        self.log(f"Found {len(year_columns)} year columns: {year_columns[0][1]} to {year_columns[-1][1]}")

        # Read data rows
        # Columns: 1=Scenario, 2=Country, 3=Tech.Name, 4=Tech, 5=Parameter
        edit_instructions = []
        for row_idx in range(2, ws.max_row + 1):
            scenario = ws.cell(row_idx, 1).value
            country = ws.cell(row_idx, 2).value
            tech_name = ws.cell(row_idx, 3).value
            tech = ws.cell(row_idx, 4).value  # This is the Tech code (auto-filled from Tech.Name)
            parameter = ws.cell(row_idx, 5).value

            # Skip empty rows
            if not scenario and not country and not tech and not parameter:
                continue

            # Read year values
            year_values = {}
            for col_idx, year in year_columns:
                value = ws.cell(row_idx, col_idx).value
                if value is not None and value != "":
                    year_values[year] = value

            # Only add if we have at least one year value
            if year_values:
                edit_instructions.append({
                    'row': row_idx,
                    'scenario': str(scenario).strip() if scenario else None,
                    'country': str(country).strip() if country else None,
                    'tech_name': str(tech_name).strip() if tech_name else None,  # Keep for logging
                    'tech': str(tech).strip() if tech else None,  # This is what we'll use for matching
                    'parameter': str(parameter).strip() if parameter else None,
                    'year_values': year_values
                })

        wb.close()

        self.log(f"Found {len(edit_instructions)} rows with data to process")
        return edit_instructions

    def validate_instruction(self, instruction):
        """
        Validate an edit instruction

        Returns:
            (is_valid, error_message)
        """
        if not instruction['scenario']:
            return False, "Scenario is empty"

        if not instruction['country']:
            return False, "Country is empty"

        if not instruction['tech']:
            return False, "Tech is empty"

        if not instruction['parameter']:
            return False, "Parameter is empty"

        if not instruction['year_values']:
            return False, "No year values provided"

        # Validate scenario
        valid_scenarios = self.scenarios + ['ALL']
        if instruction['scenario'] not in valid_scenarios:
            return False, f"Invalid scenario '{instruction['scenario']}'. Must be one of: {valid_scenarios}"

        # Validate that tech contains country code (for PWR technologies: PWRTRNARGXX -> ARG is at position 6-8)
        tech = instruction['tech'].upper()
        country = instruction['country'].upper()

        if tech.startswith('PWR'):
            # For PWR technologies, country code is at positions 6-8
            if len(tech) >= 9:
                tech_country = tech[6:9]
                if tech_country != country:
                    return False, f"Tech '{instruction['tech']}' contains country code '{tech_country}', but '{country}' was specified"
            else:
                return False, f"Tech '{instruction['tech']}' has invalid format (too short for PWR technology)"
        else:
            # For non-PWR technologies, country code might be at the beginning or elsewhere
            # We'll just issue a warning but allow it
            pass

        return True, None

    def create_backup(self, file_path):
        """Create backup of file before modifying"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.with_name(f"{file_path.stem}_backup_{timestamp}{file_path.suffix}")
        shutil.copy2(file_path, backup_path)
        return backup_path

    def find_and_update_row(self, ws, tech, parameter, year_values, year_col_map):
        """
        Find the row matching tech and parameter, and update year values

        Returns:
            (success, message, values_updated)
        """
        # Search for matching row
        # Columns in Secondary Techs: 1=Tech.Id, 2=Tech, 3=Tech.Name, 5=Parameter
        target_row = None
        for row_idx in range(2, ws.max_row + 1):
            row_tech = ws.cell(row_idx, 2).value  # Column 2: Tech
            row_param = ws.cell(row_idx, 5).value  # Column 5: Parameter

            if row_tech and row_param:
                if str(row_tech).strip() == tech and str(row_param).strip() == parameter:
                    target_row = row_idx
                    break

        if not target_row:
            return False, f"No matching row found for Tech='{tech}' and Parameter='{parameter}'", 0

        # Update year values
        values_updated = 0
        for year, value in year_values.items():
            if year in year_col_map:
                col_idx = year_col_map[year]
                ws.cell(target_row, col_idx, value)
                values_updated += 1

        return True, f"Row {target_row} updated with {values_updated} year values", values_updated

    def apply_instruction_to_scenario(self, instruction, scenario):
        """
        Apply a single edit instruction to a specific scenario

        Returns:
            (success, message)
        """
        scenario_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"

        if not scenario_path.exists():
            return False, f"File not found: {scenario_path}"

        # Create backup
        backup_path = self.create_backup(scenario_path)
        self.log(f"  Backup created: {backup_path.name}", "DEBUG")

        try:
            # Open workbook
            wb = openpyxl.load_workbook(scenario_path)

            if 'Secondary Techs' not in wb.sheetnames:
                wb.close()
                return False, f"'Secondary Techs' sheet not found in {scenario}"

            ws = wb['Secondary Techs']

            # Build year column map
            headers = [cell.value for cell in ws[1]]
            year_col_map = {}
            for col_idx, header in enumerate(headers, 1):
                if header and str(header).isdigit():
                    try:
                        year = int(header)
                        if 2000 <= year <= 2100:
                            year_col_map[year] = col_idx
                    except:
                        pass

            # Apply update
            success, message, values_updated = self.find_and_update_row(
                ws,
                instruction['tech'],
                instruction['parameter'],
                instruction['year_values'],
                year_col_map
            )

            if success:
                wb.save(scenario_path)
                wb.close()
                self.changes_applied += 1
                return True, f"{scenario}: {message}"
            else:
                wb.close()
                return False, f"{scenario}: {message}"

        except Exception as e:
            return False, f"{scenario}: Error - {str(e)}"

    def apply_instruction(self, instruction):
        """
        Apply a single edit instruction to the appropriate scenario(s)
        """
        row_num = instruction['row']
        self.log(f"\nProcessing Row {row_num}:", "INFO")
        self.log(f"  Scenario: {instruction['scenario']}", "DEBUG")
        self.log(f"  Country: {instruction['country']}", "DEBUG")
        self.log(f"  Tech.Name: {instruction['tech_name']}", "DEBUG")
        self.log(f"  Tech: {instruction['tech']}", "DEBUG")
        self.log(f"  Parameter: {instruction['parameter']}", "DEBUG")
        self.log(f"  Year values: {len(instruction['year_values'])} years", "DEBUG")

        # Validate
        is_valid, error_msg = self.validate_instruction(instruction)
        if not is_valid:
            self.log(f"  ✗ FAILED: {error_msg}", "ERROR")
            self.rows_failed += 1
            return

        # Determine which scenarios to apply to
        if instruction['scenario'] == 'ALL':
            target_scenarios = self.scenarios
            self.log(f"  Applying to ALL scenarios", "DEBUG")
        else:
            target_scenarios = [instruction['scenario']]

        # Apply to each target scenario
        all_success = True
        for scenario in target_scenarios:
            success, message = self.apply_instruction_to_scenario(instruction, scenario)

            if success:
                self.log(f"  ✓ {message}", "SUCCESS")
            else:
                self.log(f"  ✗ {message}", "ERROR")
                all_success = False

        if not all_success:
            self.rows_failed += 1

    def save_log(self, log_path):
        """Save log to file"""
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.log_lines))

    def run(self):
        """Main execution"""
        self.log("=" * 80)
        self.log("SECONDARY TECHS UPDATER")
        self.log("=" * 80)
        self.log(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log(f"Editor file: {self.editor_path}")
        self.log("")

        try:
            # Read editor file
            instructions = self.read_editor_file()

            if not instructions:
                self.log("No data found in editor file. Nothing to update.", "WARNING")
                return 0

            # Apply each instruction
            for instruction in instructions:
                self.apply_instruction(instruction)

            # Summary
            self.log("")
            self.log("=" * 80)
            self.log("SUMMARY")
            self.log("=" * 80)
            self.log(f"Total rows processed: {len(instructions)}")
            self.log(f"Changes applied: {self.changes_applied}")
            self.log(f"Rows failed: {self.rows_failed}")

            # Save log
            log_path = self.base_path / f"secondary_techs_update_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            self.save_log(log_path)
            self.log(f"\nLog saved: {log_path}")

            if self.rows_failed > 0:
                self.log("\n⚠ Some rows failed. Please review the log.", "WARNING")
                return 1
            else:
                self.log("\n✓ All changes applied successfully!", "SUCCESS")
                return 0

        except Exception as e:
            self.log(f"\nFATAL ERROR: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return 1


def main():
    try:
        # Paths
        script_dir = Path(__file__).parent
        editor_path = script_dir / "Secondary_Techs_Editor.xlsx"
        base_path = script_dir / "A1_Outputs"

        # Create updater and run
        updater = SecondaryTechsUpdater(editor_path, base_path)
        return updater.run()

    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
