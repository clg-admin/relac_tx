"""
Generate Secondary Techs Editor Template

This script reads all A-O_Parametrization.xlsx files and generates
a user-friendly Excel template for editing Secondary Techs data.

Usage:
    python t1_confection/generate_editor_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import sys
from pathlib import Path
from datetime import datetime

# OLADE country name to ISO-3 code mapping
OLADE_COUNTRY_MAPPING = {
    'Argentina': 'ARG',
    'Barbados': 'BAR',
    'Belice': 'BLZ',
    'Bolivia': 'BOL',
    'Brasil': 'BRA',
    'Chile': 'CHI',
    'Colombia': 'COL',
    'Costa Rica': 'CRC',
    'Cuba': 'CUB',
    'Ecuador': 'ECU',
    'El Salvador': 'SLV',
    'Grenada': 'GRD',
    'Guatemala': 'GTM',
    'Guyana': 'GUY',
    'Haiti': 'HTI',
    'Honduras': 'HND',
    'Jamaica': 'JAM',
    'México': 'MEX',
    'Nicaragua': 'NIC',
    'Panamá': 'PAN',
    'Paraguay': 'PRY',
    'Perú': 'PER',
    'República Dominicana': 'DOM',
    'Suriname': 'SUR',
    'Trinidad & Tobago': 'TTO',
    'Uruguay': 'URY',
    'Venezuela': 'VEN'
}

# OLADE technology to model tech code (3 chars) mapping
OLADE_TECH_MAPPING = {
    'Nuclear': 'URN',
    'Gas natural': 'CCG',
    'Carbón mineral': 'COA',
    'Hidro': 'HYD',
    'Geotermia': 'GEO',
    'Eólica': 'WON',
    'Solar': 'SPV'
    # Note: BIO is special - sum of 'Biogás' + 'Biomasa sólida'
    # Note: 'Petróleo y derivados' pending confirmation
}


def collect_data_from_all_scenarios():
    """
    Collect all unique values from all scenario files

    Returns:
        dict with scenarios, countries, tech_mapping (Tech.Name -> Tech), parameters, years
    """
    base_path = Path(__file__).parent / "A1_Outputs"

    scenarios = ["BAU", "NDC", "NDC+ELC", "NDC_NoRPO"]

    all_countries = set()
    tech_mapping = {}  # Tech.Name -> Tech code
    all_parameters = set()
    all_years = set()

    print("Collecting data from all scenarios...")
    print()

    for scenario in scenarios:
        scenario_path = base_path / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"

        if not scenario_path.exists():
            print(f"WARNING: File not found: {scenario_path}")
            continue

        print(f"Reading {scenario}...")

        wb = openpyxl.load_workbook(scenario_path, data_only=True)

        if 'Secondary Techs' not in wb.sheetnames:
            print(f"  WARNING: 'Secondary Techs' sheet not found in {scenario}")
            wb.close()
            continue

        ws = wb['Secondary Techs']

        # Get headers to find year columns
        headers = [cell.value for cell in ws[1]]

        # Find year columns (columns with numeric values representing years)
        year_col_indices = []
        for idx, header in enumerate(headers, 1):
            if header and str(header).isdigit():
                try:
                    year = int(header)
                    if 2000 <= year <= 2100:
                        all_years.add(year)
                        year_col_indices.append(idx)
                except:
                    pass

        # Collect tech mapping and parameters (skip header row)
        # Columns: 1=Tech.Id, 2=Tech, 3=Tech.Name, 5=Parameter
        for row_idx in range(2, ws.max_row + 1):
            tech_code = ws.cell(row_idx, 2).value      # Column 2: Tech
            tech_name = ws.cell(row_idx, 3).value      # Column 3: Tech.Name
            parameter = ws.cell(row_idx, 5).value      # Column 5: Parameter

            if tech_code and tech_name:
                tech_code_str = str(tech_code).strip()
                tech_name_str = str(tech_name).strip()

                # Build mapping: Tech.Name -> Tech
                tech_mapping[tech_name_str] = tech_code_str

                # Extract country code from PWR technologies only
                # Format: PWRTRNARGXX -> country code is ARG (characters 6-8)
                if tech_code_str.upper().startswith('PWR') and len(tech_code_str) >= 9:
                    country_code = tech_code_str[6:9].upper()
                    all_countries.add(country_code)

            if parameter:
                all_parameters.add(str(parameter).strip())

        wb.close()
        print(f"  Found: {len(tech_mapping)} technologies, {len(all_parameters)} parameters")

    print()
    print(f"Summary:")
    print(f"  Scenarios: {len(scenarios)}")
    print(f"  Countries: {len(all_countries)} - {sorted(all_countries)}")
    print(f"  Technologies (Tech.Name): {len(tech_mapping)}")
    print(f"  Parameters: {len(all_parameters)}")
    print(f"  Years: {len(all_years)} - {min(all_years) if all_years else 'N/A'} to {max(all_years) if all_years else 'N/A'}")
    print()

    return {
        'scenarios': sorted(scenarios),
        'countries': sorted(all_countries),
        'tech_mapping': tech_mapping,  # Tech.Name -> Tech code
        'tech_names': sorted(tech_mapping.keys()),  # List of Tech.Name for dropdown
        'parameters': sorted(all_parameters),
        'years': sorted(all_years)
    }


def create_editor_template(data, output_path):
    """
    Create the Excel editor template with dropdowns and validation

    Args:
        data: dict with scenarios, countries, technologies, parameters, years
        output_path: Path where to save the template
    """
    print("Creating Excel template...")

    wb = openpyxl.Workbook()

    # Main sheet for data entry
    ws_main = wb.active
    ws_main.title = "Editor"

    # Create hidden sheets for dropdown lists and mappings
    ws_scenarios = wb.create_sheet("_Scenarios")
    ws_countries = wb.create_sheet("_Countries")
    ws_tech_names = wb.create_sheet("_TechNames")
    ws_tech_mapping = wb.create_sheet("_TechMapping")  # Tech.Name -> Tech code
    ws_parameters = wb.create_sheet("_Parameters")

    # Hide validation sheets
    ws_scenarios.sheet_state = 'hidden'
    ws_countries.sheet_state = 'hidden'
    ws_tech_names.sheet_state = 'hidden'
    ws_tech_mapping.sheet_state = 'hidden'
    ws_parameters.sheet_state = 'hidden'

    # Populate validation sheets
    for idx, scenario in enumerate(['ALL'] + data['scenarios'], 1):
        ws_scenarios.cell(idx, 1, scenario)

    for idx, country in enumerate(data['countries'], 1):
        ws_countries.cell(idx, 1, country)

    # Populate Tech.Name list and Tech.Name -> Tech mapping
    for idx, (tech_name, tech_code) in enumerate(sorted(data['tech_mapping'].items()), 1):
        ws_tech_names.cell(idx, 1, tech_name)  # Column A: Tech.Name for dropdown
        ws_tech_mapping.cell(idx, 1, tech_name)  # Column A: Tech.Name
        ws_tech_mapping.cell(idx, 2, tech_code)  # Column B: Tech code

    for idx, param in enumerate(data['parameters'], 1):
        ws_parameters.cell(idx, 1, param)

    # Create header row in main sheet
    # Columns: Scenario, Country, Tech.Name, Tech (auto-filled), Parameter, Years...
    headers = ['Scenario', 'Country', 'Tech.Name', 'Tech', 'Parameter'] + [str(year) for year in data['years']]

    # Style for header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Set column widths
    ws_main.column_dimensions['A'].width = 15  # Scenario
    ws_main.column_dimensions['B'].width = 12  # Country
    ws_main.column_dimensions['C'].width = 40  # Tech.Name
    ws_main.column_dimensions['D'].width = 20  # Tech (auto-filled, read-only)
    ws_main.column_dimensions['E'].width = 30  # Parameter

    # Year columns (start from column F = 6)
    for col_idx in range(6, len(headers) + 1):
        ws_main.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10

    # Add data validations (dropdowns) for first 100 rows
    max_data_rows = 100

    # Scenario dropdown (column A)
    dv_scenario = DataValidation(
        type="list",
        formula1=f"=_Scenarios!$A$1:$A${len(data['scenarios']) + 1}",
        allow_blank=False
    )
    dv_scenario.error = 'Please select a valid scenario'
    dv_scenario.errorTitle = 'Invalid Scenario'
    ws_main.add_data_validation(dv_scenario)
    dv_scenario.add(f'A2:A{max_data_rows + 1}')

    # Country dropdown (column B)
    dv_country = DataValidation(
        type="list",
        formula1=f"=_Countries!$A$1:$A${len(data['countries'])}",
        allow_blank=False
    )
    dv_country.error = 'Please select a valid country'
    dv_country.errorTitle = 'Invalid Country'
    ws_main.add_data_validation(dv_country)
    dv_country.add(f'B2:B{max_data_rows + 1}')

    # Tech.Name dropdown (column C)
    dv_tech_name = DataValidation(
        type="list",
        formula1=f"=_TechNames!$A$1:$A${len(data['tech_names'])}",
        allow_blank=False
    )
    dv_tech_name.error = 'Please select a valid technology name'
    dv_tech_name.errorTitle = 'Invalid Tech.Name'
    ws_main.add_data_validation(dv_tech_name)
    dv_tech_name.add(f'C2:C{max_data_rows + 1}')

    # Column D (Tech) will be auto-filled with VLOOKUP formula
    # Add formulas to map Tech.Name to Tech code
    for row_idx in range(2, max_data_rows + 2):
        formula = f'=IFERROR(VLOOKUP(C{row_idx},_TechMapping!$A:$B,2,FALSE),"")'
        ws_main.cell(row_idx, 4, formula)  # Column D

    # Protect column D (Tech) from manual editing
    for row_idx in range(2, max_data_rows + 2):
        ws_main.cell(row_idx, 4).protection = Protection(locked=True)

    # Parameter dropdown (column E)
    dv_parameter = DataValidation(
        type="list",
        formula1=f"=_Parameters!$A$1:$A${len(data['parameters'])}",
        allow_blank=False
    )
    dv_parameter.error = 'Please select a valid parameter'
    dv_parameter.errorTitle = 'Invalid Parameter'
    ws_main.add_data_validation(dv_parameter)
    dv_parameter.add(f'E2:E{max_data_rows + 1}')

    # Freeze panes (freeze first row and first 5 columns)
    ws_main.freeze_panes = 'F2'

    # Add instructions in a separate sheet (first sheet - index 0)
    ws_instructions = wb.create_sheet("Instructions", 0)
    ws_instructions.column_dimensions['A'].width = 80

    # Add OLADE Configuration sheet (second sheet - index 1)
    ws_olade = wb.create_sheet("OLADE_Config", 1)
    ws_olade.column_dimensions['A'].width = 80
    ws_olade.column_dimensions['B'].width = 20

    # Style for configuration sheet
    header_fill_config = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    header_font_config = Font(bold=True, color="FFFFFF", size=12)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    cell = ws_olade.cell(1, 1, "OLADE CONFIGURATION")
    cell.font = Font(size=14, bold=True, color="366092")
    ws_olade.merge_cells('A1:B1')

    # Instructions
    ws_olade.cell(2, 1, "Fill in the configuration below to automatically populate ResidualCapacity from OLADE data")
    ws_olade.merge_cells('A2:B2')
    ws_olade.cell(2, 1).font = Font(italic=True)

    # Configuration headers
    ws_olade.cell(4, 1, "Configuration Parameter").fill = header_fill_config
    ws_olade.cell(4, 1).font = header_font_config
    ws_olade.cell(4, 1).border = border_style
    ws_olade.cell(4, 2, "Value").fill = header_fill_config
    ws_olade.cell(4, 2).font = header_font_config
    ws_olade.cell(4, 2).border = border_style

    # ResidualCapacitiesFromOLADE
    ws_olade.cell(5, 1, "ResidualCapacitiesFromOLADE").border = border_style
    ws_olade.cell(5, 2, "NO").border = border_style
    dv_yes_no = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    ws_olade.add_data_validation(dv_yes_no)
    dv_yes_no.add('B5')

    # CapacityFactorGrowth
    ws_olade.cell(6, 1, "CapacityFactorGrowth (%)").border = border_style
    ws_olade.cell(6, 2, 5.0).border = border_style
    ws_olade.cell(6, 2).number_format = '0.00'

    # GrowthType
    ws_olade.cell(7, 1, "GrowthType").border = border_style
    ws_olade.cell(7, 2, "Compound").border = border_style
    dv_growth = DataValidation(type="list", formula1='"Compound,Simple"', allow_blank=False)
    ws_olade.add_data_validation(dv_growth)
    dv_growth.add('B7')

    # Add descriptions
    ws_olade.cell(9, 1, "DESCRIPTIONS:")
    ws_olade.cell(9, 1).font = Font(bold=True, size=11)
    ws_olade.merge_cells('A9:B9')

    descriptions = [
        ("ResidualCapacitiesFromOLADE:", "Set to YES to automatically populate ResidualCapacity parameter from OLADE data. This applies only to PWR technologies (power generation). OLADE data is automatically converted from MW to GW."),
        ("CapacityFactorGrowth (%):", "Growth rate to apply for years before and after the OLADE reference year (2023). Example: 5.0 means 5% growth per year."),
        ("GrowthType:", "Compound: Exponential growth (e.g., 2024 = 2023 × (1+rate)^1). Simple: Linear growth (e.g., 2024 = 2023 + 2023×rate)."),
    ]

    current_row = 11
    for label, desc in descriptions:
        # Label in bold
        ws_olade.cell(current_row, 1, label)
        ws_olade.cell(current_row, 1).font = Font(bold=True, size=10)
        ws_olade.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

        # Description
        ws_olade.cell(current_row, 1, desc)
        ws_olade.cell(current_row, 1).font = Font(size=9)
        ws_olade.cell(current_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws_olade.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

        # Blank line
        current_row += 1

    # Populate instructions sheet
    instructions = [
        ["Secondary Techs Editor - Instructions", ""],
        ["", ""],
        ["This Excel file allows you to edit Secondary Techs data easily.", ""],
        ["", ""],
        ["HOW TO USE:", ""],
        ["1. (Optional) Configure OLADE settings in 'OLADE_Config' sheet", ""],
        ["2. Go to the 'Editor' sheet", ""],
        ["3. Fill in each row with:", ""],
        ["   - Scenario: Select BAU, NDC, NDC+ELC, NDC_NoRPO, or ALL (applies to all scenarios)", ""],
        ["   - Country: Select the country code (ARG, BOL, CHI, COL, ECU, GUA, etc.)", ""],
        ["   - Tech.Name: Select the descriptive technology name from the dropdown", ""],
        ["   - Tech: This column is auto-filled based on Tech.Name (READ-ONLY)", ""],
        ["   - Parameter: Select the parameter to modify", ""],
        ["   - Year values: Enter numeric values for each year (leave empty to keep current value)", ""],
        ["", ""],
        ["4. Save and close this file", ""],
        ["5. Run: python t1_confection/D2_update_secondary_techs.py", ""],
        ["", ""],
        ["OLADE INTEGRATION:", ""],
        ["- If ResidualCapacitiesFromOLADE = YES in OLADE_Config sheet:", ""],
        ["  * The script will automatically populate ResidualCapacity for PWR technologies", ""],
        ["  * Data comes from 'Capacidad instalada por fuente - Anual - OLADE.xlsx'", ""],
        ["  * OLADE data takes priority over manual Editor entries for ResidualCapacity", ""],
        ["  * Growth rate is applied for years before/after the reference year (2023)", ""],
        ["", ""],
        ["IMPORTANT NOTES:", ""],
        ["- You can add as many rows as needed", ""],
        ["- The 'Tech' column is automatically filled when you select a 'Tech.Name'", ""],
        ["- Empty year cells will NOT modify the current value in the destination file", ""],
        ["- Scenario 'ALL' will apply changes to all 4 scenario files", ""],
        ["- A backup will be created automatically before making changes", ""],
        ["- Check the log file after running the update script", ""],
        ["", ""],
        [f"Template generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ""],
    ]

    for row_idx, row_data in enumerate(instructions, 1):
        cell = ws_instructions.cell(row_idx, 1, row_data[0])
        if row_idx == 1:
            cell.font = Font(size=16, bold=True, color="366092")
        elif row_data[0].startswith("HOW TO USE:") or row_data[0].startswith("IMPORTANT NOTES:") or row_data[0].startswith("OLADE INTEGRATION:"):
            cell.font = Font(size=12, bold=True)

        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save the workbook
    wb.save(output_path)
    print(f"Template saved: {output_path}")
    print()


def main():
    try:
        print("=" * 80)
        print("SECONDARY TECHS EDITOR - TEMPLATE GENERATOR")
        print("=" * 80)
        print()

        # Collect data from all scenarios
        data = collect_data_from_all_scenarios()

        if not data['tech_mapping'] or not data['parameters'] or not data['years']:
            print("ERROR: Could not collect enough data from scenario files")
            print("Please check that A-O_Parametrization.xlsx files exist and contain Secondary Techs sheet")
            return 1

        # Create template
        output_path = Path(__file__).parent / "Secondary_Techs_Editor.xlsx"
        create_editor_template(data, output_path)

        print("=" * 80)
        print("TEMPLATE GENERATION COMPLETE")
        print("=" * 80)
        print()
        print("Next steps:")
        print("1. Open t1_confection/Secondary_Techs_Editor.xlsx")
        print("2. Fill in the 'Editor' sheet with your changes")
        print("3. Save and close the file")
        print("4. Run: python t1_confection/update_secondary_techs.py")
        print()

        return 0

    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
