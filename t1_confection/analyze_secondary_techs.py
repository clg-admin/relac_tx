"""
Analyze Secondary Techs sheet structure
Run this to understand the structure before implementing the editor
"""
import openpyxl
import sys
from pathlib import Path

def analyze_secondary_techs():
    # Path to one of the A-O_Parametrization files
    base_path = Path(__file__).parent / "A1_Outputs" / "A1_Outputs_BAU"
    file_path = base_path / "A-O_Parametrization.xlsx"

    if not file_path.exists():
        print(f"ERROR: File not found: {file_path}")
        return

    print("=" * 80)
    print("ANALYZING SECONDARY TECHS STRUCTURE")
    print("=" * 80)
    print(f"File: {file_path}")
    print()

    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if 'Secondary Techs' not in wb.sheetnames:
        print("ERROR: 'Secondary Techs' sheet not found!")
        print(f"Available sheets: {wb.sheetnames}")
        wb.close()
        return

    ws = wb['Secondary Techs']

    # Get dimensions
    print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
    print()

    # Analyze headers (first row)
    print("=" * 80)
    print("HEADERS (First Row)")
    print("=" * 80)
    headers = []
    for col_idx, cell in enumerate(ws[1], 1):
        value = cell.value
        headers.append(value)
        if col_idx <= 10 or (value and str(value).isdigit()):  # Show first 10 or year columns
            print(f"Col {col_idx:3d}: {value}")

    # Identify year columns
    year_cols = []
    for idx, header in enumerate(headers, 1):
        if header and str(header).isdigit():
            try:
                year = int(header)
                if 2000 <= year <= 2100:
                    year_cols.append((idx, year))
            except:
                pass

    print()
    print(f"Year columns found: {len(year_cols)}")
    if year_cols:
        print(f"  First year: {year_cols[0][1]} (col {year_cols[0][0]})")
        print(f"  Last year: {year_cols[-1][1]} (col {year_cols[-1][0]})")
    print()

    # Analyze first 20 data rows
    print("=" * 80)
    print("SAMPLE DATA (First 20 rows)")
    print("=" * 80)

    for row_idx in range(2, min(22, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(11, ws.max_column + 1)):  # First 10 columns
            cell = ws.cell(row_idx, col_idx)
            row_data.append(str(cell.value) if cell.value else "")

        print(f"Row {row_idx:3d}: {' | '.join(row_data[:10])}")

    print()

    # Extract unique values from key columns
    print("=" * 80)
    print("EXTRACTING UNIQUE VALUES")
    print("=" * 80)

    # Assuming structure based on common patterns:
    # Column 1 might be: REGION, Country, or Technology
    # Let's check first few columns

    col1_values = set()
    col2_values = set()
    col3_values = set()

    for row_idx in range(2, min(500, ws.max_row + 1)):
        val1 = ws.cell(row_idx, 1).value
        val2 = ws.cell(row_idx, 2).value
        val3 = ws.cell(row_idx, 3).value

        if val1:
            col1_values.add(str(val1))
        if val2:
            col2_values.add(str(val2))
        if val3:
            col3_values.add(str(val3))

    print(f"\nColumn 1 - Unique values ({len(col1_values)}):")
    for val in sorted(list(col1_values))[:20]:
        print(f"  - {val}")
    if len(col1_values) > 20:
        print(f"  ... and {len(col1_values) - 20} more")

    print(f"\nColumn 2 - Unique values ({len(col2_values)}):")
    for val in sorted(list(col2_values))[:20]:
        print(f"  - {val}")
    if len(col2_values) > 20:
        print(f"  ... and {len(col2_values) - 20} more")

    print(f"\nColumn 3 - Unique values ({len(col3_values)}):")
    for val in sorted(list(col3_values))[:20]:
        print(f"  - {val}")
    if len(col3_values) > 20:
        print(f"  ... and {len(col3_values) - 20} more")

    # Extract countries from technologies (assuming they start with ISO-3 code)
    print()
    print("=" * 80)
    print("EXTRACTING COUNTRIES FROM TECHNOLOGIES")
    print("=" * 80)

    countries = set()
    technologies = set()

    # Try to find technology column
    tech_col_idx = None
    for idx, header in enumerate(headers, 1):
        if header and 'TECH' in str(header).upper():
            tech_col_idx = idx
            break

    if not tech_col_idx:
        # Try column 2 as technology column
        tech_col_idx = 2

    print(f"Using column {tech_col_idx} as TECHNOLOGY column")
    print()

    for row_idx in range(2, min(500, ws.max_row + 1)):
        tech = ws.cell(row_idx, tech_col_idx).value
        if tech:
            tech_str = str(tech)
            technologies.add(tech_str)
            # Extract first 3 characters as potential country code
            if len(tech_str) >= 3:
                country_code = tech_str[:3].upper()
                countries.add(country_code)

    print(f"Unique countries found ({len(countries)}):")
    for country in sorted(countries):
        print(f"  - {country}")

    print(f"\nSample technologies ({min(20, len(technologies))}):")
    for tech in sorted(list(technologies))[:20]:
        print(f"  - {tech}")
    if len(technologies) > 20:
        print(f"  ... and {len(technologies) - 20} more")

    wb.close()
    print()
    print("=" * 80)
    print("ANALYSIS COMPLETE")
    print("=" * 80)

if __name__ == "__main__":
    try:
        analyze_secondary_techs()
    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
