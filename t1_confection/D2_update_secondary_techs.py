"""
Update Secondary Techs from Editor Template

This script reads the Secondary_Techs_Editor.xlsx file and applies
the changes to the corresponding A-O_Parametrization.xlsx files.

Usage:
    python t1_confection/D2_update_secondary_techs.py
"""
import openpyxl
import sys
from pathlib import Path
from datetime import datetime
import shutil

# OLADE country name to ISO-3 code mapping
OLADE_COUNTRY_MAPPING = {
    'Argentina': 'ARG',
    'Barbados': 'BAR',
    'Belice': 'BLZ',
    'Bolivia': 'BOL',
    'Brasil': 'BRA',
    'Chile': 'CHI',
    'Colombia': 'COL',
    'Costa Rica': 'CRC',
    'Cuba': 'CUB',
    'Ecuador': 'ECU',
    'El Salvador': 'SLV',
    'Grenada': 'GRD',
    'Guatemala': 'GTM',
    'Guyana': 'GUY',
    'Haiti': 'HTI',
    'Honduras': 'HND',
    'Jamaica': 'JAM',
    'México': 'MEX',
    'Nicaragua': 'NIC',
    'Panamá': 'PAN',
    'Paraguay': 'PRY',
    'Perú': 'PER',
    'República Dominicana': 'DOM',
    'Suriname': 'SUR',
    'Trinidad & Tobago': 'TTO',
    'Uruguay': 'URY',
    'Venezuela': 'VEN'
}

# OLADE technology to model tech code (3 chars) mapping
OLADE_TECH_MAPPING = {
    'Nuclear': 'URN',
    'Gas natural': 'CCG',
    'Carbón mineral': 'COA',
    'Hidro': 'HYD',
    'Geotermia': 'GEO',
    'Eólica': 'WON',
    'Solar': 'SPV'
    # Note: BIO is special - sum of 'Biogás' + 'Biomasa sólida'
    # Note: 'Petróleo y derivados' pending confirmation
}


def read_olade_config(editor_path):
    """
    Read OLADE configuration from the editor Excel file

    Returns:
        dict with config: {enabled, growth_rate, growth_type}
    """
    wb = openpyxl.load_workbook(editor_path, data_only=True)

    if 'OLADE_Config' not in wb.sheetnames:
        wb.close()
        return {'enabled': False, 'growth_rate': 0.0, 'growth_type': 'Compound'}

    ws = wb['OLADE_Config']

    # Read configuration values (B5, B6, B7)
    enabled = str(ws['B5'].value).upper() == 'YES' if ws['B5'].value else False
    growth_rate = float(ws['B6'].value) if ws['B6'].value else 5.0
    growth_type = str(ws['B7'].value) if ws['B7'].value else 'Compound'

    wb.close()

    return {
        'enabled': enabled,
        'growth_rate': growth_rate,
        'growth_type': growth_type
    }


def read_olade_data(olade_file_path):
    """
    Read OLADE capacity data from Excel file

    Note: OLADE data is in MW, but is converted to GW for the model (1 GW = 1000 MW)

    Returns:
        dict: {
            'reference_year': int,
            'data': {
                country_iso3: {
                    tech_code: capacity_gw
                }
            }
        }
    """
    if not olade_file_path.exists():
        raise FileNotFoundError(f"OLADE file not found: {olade_file_path}")

    wb = openpyxl.load_workbook(olade_file_path, data_only=True)
    ws = wb['1.2023']

    # Extract reference year from A4
    ref_year_cell = ws['A4'].value
    ref_year = 2023  # Default
    if ref_year_cell and str(ref_year_cell).startswith('2023'):
        ref_year = 2023

    # Get country columns from row 5 (starting at column 3)
    country_columns = {}
    for col_idx in range(3, ws.max_column + 1):
        country_name = ws.cell(5, col_idx).value
        if country_name and str(country_name) in OLADE_COUNTRY_MAPPING:
            iso3_code = OLADE_COUNTRY_MAPPING[str(country_name)]
            country_columns[col_idx] = iso3_code

    # Read technology data (rows 6-20)
    data = {}

    # Process each technology
    for row_idx in range(6, 21):
        tech_name = ws.cell(row_idx, 1).value
        if not tech_name:
            continue

        tech_name_str = str(tech_name).strip()

        # Skip non-applicable technologies
        if tech_name_str in ['Térmica no renovable (combustión)', 'Otras fuentes',
                             'Térmica renovable (combustión)', 'Fuentes renovable (no combustión)',
                             'Biocombustibles líquidos', 'Total']:
            continue

        # Map OLADE tech name to model tech code
        tech_code = None
        if tech_name_str in OLADE_TECH_MAPPING:
            tech_code = OLADE_TECH_MAPPING[tech_name_str]

        # Special handling for BIO (sum of Biogás and Biomasa sólida)
        if tech_name_str == 'Biogás':
            tech_code = 'BIO'
        elif tech_name_str == 'Biomasa sólida':
            tech_code = 'BIO'

        if not tech_code:
            continue

        # Read capacity values for each country
        for col_idx, country_iso3 in country_columns.items():
            capacity = ws.cell(row_idx, col_idx).value

            if capacity is not None and capacity != '':
                try:
                    capacity_mw = float(capacity)

                    # Convert from MW to GW (1 GW = 1000 MW)
                    capacity_gw = capacity_mw / 1000.0

                    # Initialize country if not exists
                    if country_iso3 not in data:
                        data[country_iso3] = {}

                    # For BIO, sum Biogás + Biomasa sólida
                    if tech_code == 'BIO':
                        if tech_code in data[country_iso3]:
                            data[country_iso3][tech_code] += capacity_gw
                        else:
                            data[country_iso3][tech_code] = capacity_gw
                    else:
                        data[country_iso3][tech_code] = capacity_gw

                except ValueError:
                    pass

    wb.close()

    return {
        'reference_year': ref_year,
        'data': data
    }


def calculate_capacity_for_year(base_capacity, base_year, target_year, growth_rate, growth_type):
    """
    Calculate capacity for a target year based on growth rate

    Args:
        base_capacity: Capacity at base year (GW)
        base_year: Reference year
        target_year: Year to calculate
        growth_rate: Growth rate as percentage (e.g., 5.0 for 5%)
        growth_type: 'Compound' or 'Simple'

    Returns:
        Calculated capacity (GW)
    """
    years_diff = target_year - base_year
    rate = growth_rate / 100.0

    if years_diff == 0:
        return base_capacity

    if growth_type == 'Compound':
        # Exponential growth: capacity = base * (1 + rate)^years_diff
        return base_capacity * ((1 + rate) ** years_diff)
    else:  # Simple
        # Linear growth: capacity = base + (base * rate * years_diff)
        return base_capacity + (base_capacity * rate * years_diff)


class SecondaryTechsUpdater:
    def __init__(self, editor_path, base_path, olade_file_path=None):
        self.editor_path = editor_path
        self.base_path = base_path
        self.olade_file_path = olade_file_path
        self.scenarios = ["BAU", "NDC", "NDC+ELC", "NDC_NoRPO"]
        self.log_lines = []
        self.changes_applied = 0
        self.rows_failed = 0
        self.olade_config = None
        self.olade_data = None

    def log(self, message, level="INFO"):
        """Add message to log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_line = f"[{timestamp}] {level}: {message}"
        self.log_lines.append(log_line)
        print(log_line)

    def read_editor_file(self):
        """
        Read and parse the editor Excel file

        Returns:
            list of dicts with editing instructions
        """
        self.log("Reading editor file...")

        if not self.editor_path.exists():
            raise FileNotFoundError(f"Editor file not found: {self.editor_path}")

        wb = openpyxl.load_workbook(self.editor_path, data_only=True)

        if 'Editor' not in wb.sheetnames:
            raise ValueError("'Editor' sheet not found in template file")

        ws = wb['Editor']

        # Read header to get year columns
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                break

        # Identify year columns (after Scenario, Country, Tech.Name, Tech, Parameter)
        # Years start from column 6 (F)
        year_columns = []
        for idx, header in enumerate(headers[5:], 6):  # Start from column 6 (F)
            if header.isdigit():
                year_columns.append((idx, int(header)))

        self.log(f"Found {len(year_columns)} year columns: {year_columns[0][1]} to {year_columns[-1][1]}")

        # Read data rows
        # Columns: 1=Scenario, 2=Country, 3=Tech.Name, 4=Tech, 5=Parameter
        edit_instructions = []
        for row_idx in range(2, ws.max_row + 1):
            scenario = ws.cell(row_idx, 1).value
            country = ws.cell(row_idx, 2).value
            tech_name = ws.cell(row_idx, 3).value
            tech = ws.cell(row_idx, 4).value  # This is the Tech code (auto-filled from Tech.Name)
            parameter = ws.cell(row_idx, 5).value

            # Skip empty rows
            if not scenario and not country and not tech and not parameter:
                continue

            # Read year values
            year_values = {}
            for col_idx, year in year_columns:
                value = ws.cell(row_idx, col_idx).value
                if value is not None and value != "":
                    year_values[year] = value

            # Only add if we have at least one year value
            if year_values:
                edit_instructions.append({
                    'row': row_idx,
                    'scenario': str(scenario).strip() if scenario else None,
                    'country': str(country).strip() if country else None,
                    'tech_name': str(tech_name).strip() if tech_name else None,  # Keep for logging
                    'tech': str(tech).strip() if tech else None,  # This is what we'll use for matching
                    'parameter': str(parameter).strip() if parameter else None,
                    'year_values': year_values
                })

        wb.close()

        self.log(f"Found {len(edit_instructions)} rows with data to process")
        return edit_instructions

    def validate_instruction(self, instruction):
        """
        Validate an edit instruction

        Returns:
            (is_valid, error_message)
        """
        if not instruction['scenario']:
            return False, "Scenario is empty"

        if not instruction['country']:
            return False, "Country is empty"

        if not instruction['tech']:
            return False, "Tech is empty"

        if not instruction['parameter']:
            return False, "Parameter is empty"

        if not instruction['year_values']:
            return False, "No year values provided"

        # Validate scenario
        valid_scenarios = self.scenarios + ['ALL']
        if instruction['scenario'] not in valid_scenarios:
            return False, f"Invalid scenario '{instruction['scenario']}'. Must be one of: {valid_scenarios}"

        # Skip tech validation for OLADE instructions (they use 3-char codes)
        if instruction.get('is_olade'):
            return True, None

        # Validate that tech contains country code (for PWR technologies: PWRTRNARGXX -> ARG is at position 6-8)
        tech = instruction['tech'].upper()
        country = instruction['country'].upper()

        if tech.startswith('PWR'):
            # For PWR technologies, country code is at positions 6-8
            if len(tech) >= 9:
                tech_country = tech[6:9]
                if tech_country != country:
                    return False, f"Tech '{instruction['tech']}' contains country code '{tech_country}', but '{country}' was specified"
            else:
                return False, f"Tech '{instruction['tech']}' has invalid format (too short for PWR technology)"
        else:
            # For non-PWR technologies, country code might be at the beginning or elsewhere
            # We'll just issue a warning but allow it
            pass

        return True, None

    def create_backup(self, file_path):
        """Create backup of file before modifying"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.with_name(f"{file_path.stem}_backup_{timestamp}{file_path.suffix}")
        shutil.copy2(file_path, backup_path)
        return backup_path

    def find_and_update_row(self, ws, tech, parameter, year_values, year_col_map, projection_mode_col, is_olade=False, country=None):
        """
        Find the row matching tech and parameter, and update year values

        Args:
            is_olade: If True, tech is a 3-char code and we need to match PWR technologies by first 9 chars
            country: Country code (needed for OLADE matching)

        Returns:
            (success, message, values_updated)
        """
        # Search for matching row
        # Columns in Secondary Techs: 1=Tech.Id, 2=Tech, 3=Tech.Name, 5=Parameter
        target_rows = []

        for row_idx in range(2, ws.max_row + 1):
            row_tech = ws.cell(row_idx, 2).value  # Column 2: Tech
            row_param = ws.cell(row_idx, 5).value  # Column 5: Parameter

            if not row_tech or not row_param:
                continue

            row_tech_str = str(row_tech).strip()
            row_param_str = str(row_param).strip()

            # Check parameter match
            if row_param_str != parameter:
                continue

            # Check tech match
            if is_olade:
                # For OLADE: match PWR technologies by first 9 chars (PWR + 3 chars + country code)
                # Example: Looking for tech_code='URN' country='ARG' should match 'PWRURNARGXX'
                if row_tech_str.upper().startswith('PWR') and len(row_tech_str) >= 9:
                    # Extract positions 4-6 (tech type) and 7-9 (country)
                    row_tech_type = row_tech_str[3:6].upper()
                    row_country = row_tech_str[6:9].upper()

                    if row_tech_type == tech.upper() and row_country == country.upper():
                        target_rows.append(row_idx)
            else:
                # Exact match for manual instructions
                if row_tech_str == tech:
                    target_rows.append(row_idx)

        if not target_rows:
            tech_desc = f"Tech type='{tech}' Country='{country}'" if is_olade else f"Tech='{tech}'"
            return False, f"No matching row found for {tech_desc} and Parameter='{parameter}'", 0

        # Update year values for all matching rows
        total_values_updated = 0
        rows_updated = []

        for target_row in target_rows:
            values_updated = 0
            for year, value in year_values.items():
                if year in year_col_map:
                    col_idx = year_col_map[year]
                    ws.cell(target_row, col_idx, value)
                    values_updated += 1

            # Set Projection.Mode to "User defined" if column exists
            if projection_mode_col:
                current_value = ws.cell(target_row, projection_mode_col).value
                if current_value != "User defined":
                    ws.cell(target_row, projection_mode_col, "User defined")

            total_values_updated += values_updated
            rows_updated.append(target_row)

        if len(rows_updated) > 1:
            return True, f"Rows {rows_updated} updated with {total_values_updated} total year values", total_values_updated
        else:
            return True, f"Row {rows_updated[0]} updated with {total_values_updated} year values", total_values_updated

    def apply_instruction_to_scenario(self, instruction, scenario):
        """
        Apply a single edit instruction to a specific scenario

        Returns:
            (success, message)
        """
        scenario_path = self.base_path / f"A1_Outputs_{scenario}" / "A-O_Parametrization.xlsx"

        if not scenario_path.exists():
            return False, f"File not found: {scenario_path}"

        # Create backup
        backup_path = self.create_backup(scenario_path)
        self.log(f"  Backup created: {backup_path.name}", "DEBUG")

        try:
            # Open workbook
            wb = openpyxl.load_workbook(scenario_path)

            if 'Secondary Techs' not in wb.sheetnames:
                wb.close()
                return False, f"'Secondary Techs' sheet not found in {scenario}"

            ws = wb['Secondary Techs']

            # Build year column map and find Projection.Mode column
            headers = [cell.value for cell in ws[1]]
            year_col_map = {}
            projection_mode_col = None

            for col_idx, header in enumerate(headers, 1):
                if header:
                    # Check for year columns
                    if str(header).isdigit():
                        try:
                            year = int(header)
                            if 2000 <= year <= 2100:
                                year_col_map[year] = col_idx
                        except:
                            pass
                    # Check for Projection.Mode column
                    elif str(header).strip() == "Projection.Mode":
                        projection_mode_col = col_idx

            # Apply update
            is_olade = instruction.get('is_olade', False)
            country = instruction.get('country')

            success, message, values_updated = self.find_and_update_row(
                ws,
                instruction['tech'],
                instruction['parameter'],
                instruction['year_values'],
                year_col_map,
                projection_mode_col,
                is_olade=is_olade,
                country=country
            )

            if success:
                wb.save(scenario_path)
                wb.close()
                self.changes_applied += 1
                return True, f"{scenario}: {message}"
            else:
                wb.close()
                return False, f"{scenario}: {message}"

        except Exception as e:
            return False, f"{scenario}: Error - {str(e)}"

    def apply_instruction(self, instruction):
        """
        Apply a single edit instruction to the appropriate scenario(s)
        """
        row_num = instruction['row']
        self.log(f"\nProcessing Row {row_num}:", "INFO")
        self.log(f"  Scenario: {instruction['scenario']}", "DEBUG")
        self.log(f"  Country: {instruction['country']}", "DEBUG")
        self.log(f"  Tech.Name: {instruction['tech_name']}", "DEBUG")
        self.log(f"  Tech: {instruction['tech']}", "DEBUG")
        self.log(f"  Parameter: {instruction['parameter']}", "DEBUG")
        self.log(f"  Year values: {len(instruction['year_values'])} years", "DEBUG")

        # Validate
        is_valid, error_msg = self.validate_instruction(instruction)
        if not is_valid:
            self.log(f"  ✗ FAILED: {error_msg}", "ERROR")
            self.rows_failed += 1
            return

        # Determine which scenarios to apply to
        if instruction['scenario'] == 'ALL':
            target_scenarios = self.scenarios
            self.log(f"  Applying to ALL scenarios", "DEBUG")
        else:
            target_scenarios = [instruction['scenario']]

        # Apply to each target scenario
        all_success = True
        for scenario in target_scenarios:
            success, message = self.apply_instruction_to_scenario(instruction, scenario)

            if success:
                self.log(f"  ✓ {message}", "SUCCESS")
            else:
                self.log(f"  ✗ {message}", "ERROR")
                all_success = False

        if not all_success:
            self.rows_failed += 1

    def generate_olade_instructions(self, all_years):
        """
        Generate instructions from OLADE data

        Args:
            all_years: list of years from Secondary Techs

        Returns:
            list of instruction dicts
        """
        if not self.olade_config['enabled']:
            return []

        self.log("")
        self.log("=" * 80)
        self.log("PROCESSING OLADE DATA")
        self.log("=" * 80)
        self.log(f"Growth rate: {self.olade_config['growth_rate']}%")
        self.log(f"Growth type: {self.olade_config['growth_type']}")
        self.log(f"Reference year: {self.olade_data['reference_year']}")
        self.log("")

        instructions = []
        ref_year = self.olade_data['reference_year']

        # Generate instructions for each country and technology
        for country_iso3, techs in self.olade_data['data'].items():
            for tech_code, base_capacity in techs.items():
                # Calculate capacity for each year
                year_values = {}
                for year in all_years:
                    capacity = calculate_capacity_for_year(
                        base_capacity,
                        ref_year,
                        year,
                        self.olade_config['growth_rate'],
                        self.olade_config['growth_type']
                    )
                    year_values[year] = round(capacity, 2)

                # Create instruction for each scenario
                for scenario in self.scenarios:
                    instruction = {
                        'row': 'OLADE',
                        'scenario': scenario,
                        'country': country_iso3,
                        'tech_name': f'PWR-{tech_code}',
                        'tech': tech_code,  # This will be used to match PWR technologies
                        'parameter': 'ResidualCapacity',
                        'year_values': year_values,
                        'is_olade': True
                    }
                    instructions.append(instruction)

        self.log(f"Generated {len(instructions)} OLADE instructions")
        self.log("")

        return instructions

    def save_log(self, log_path):
        """Save log to file"""
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(self.log_lines))

    def run(self):
        """Main execution"""
        self.log("=" * 80)
        self.log("SECONDARY TECHS UPDATER")
        self.log("=" * 80)
        self.log(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.log(f"Editor file: {self.editor_path}")
        self.log("")

        try:
            # Read OLADE configuration
            self.olade_config = read_olade_config(self.editor_path)

            if self.olade_config['enabled']:
                self.log("OLADE integration: ENABLED")
                if self.olade_file_path and self.olade_file_path.exists():
                    self.log(f"OLADE file: {self.olade_file_path}")
                    try:
                        self.olade_data = read_olade_data(self.olade_file_path)
                        self.log(f"OLADE data loaded: {len(self.olade_data['data'])} countries")
                    except Exception as e:
                        self.log(f"ERROR loading OLADE data: {e}", "ERROR")
                        self.log("Continuing without OLADE data...", "WARNING")
                        self.olade_config['enabled'] = False
                else:
                    self.log(f"WARNING: OLADE file not found: {self.olade_file_path}", "WARNING")
                    self.log("Continuing without OLADE data...", "WARNING")
                    self.olade_config['enabled'] = False
            else:
                self.log("OLADE integration: DISABLED")

            self.log("")

            # Read editor file
            instructions = self.read_editor_file()

            # Get all years from first scenario to determine year range
            all_years = set()
            scenario_path = self.base_path / "A1_Outputs_BAU" / "A-O_Parametrization.xlsx"
            if scenario_path.exists():
                wb = openpyxl.load_workbook(scenario_path, data_only=True)
                if 'Secondary Techs' in wb.sheetnames:
                    ws = wb['Secondary Techs']
                    headers = [cell.value for cell in ws[1]]
                    for header in headers:
                        if header and str(header).isdigit():
                            try:
                                year = int(header)
                                if 2000 <= year <= 2100:
                                    all_years.add(year)
                            except:
                                pass
                wb.close()

            # Generate OLADE instructions if enabled
            olade_instructions = []
            if self.olade_config['enabled'] and self.olade_data:
                olade_instructions = self.generate_olade_instructions(sorted(all_years))

            # Combine instructions: OLADE takes priority for ResidualCapacity
            # Filter out manual ResidualCapacity instructions for PWR techs if OLADE is enabled
            if olade_instructions:
                filtered_instructions = []
                for instr in instructions:
                    # Check if this is a ResidualCapacity instruction for a PWR tech
                    if (instr.get('parameter') == 'ResidualCapacity' and
                        instr.get('tech') and str(instr.get('tech')).upper().startswith('PWR')):
                        # Skip it - OLADE will handle it
                        self.log(f"Skipping manual ResidualCapacity for {instr['tech']} - using OLADE data", "DEBUG")
                        continue
                    filtered_instructions.append(instr)
                instructions = filtered_instructions + olade_instructions
            else:
                # No OLADE, use manual instructions as-is
                pass

            if not instructions:
                self.log("No data found to process. Nothing to update.", "WARNING")
                return 0

            # Apply each instruction
            for instruction in instructions:
                self.apply_instruction(instruction)

            # Summary
            self.log("")
            self.log("=" * 80)
            self.log("SUMMARY")
            self.log("=" * 80)
            self.log(f"Total instructions processed: {len(instructions)}")
            self.log(f"Changes applied: {self.changes_applied}")
            self.log(f"Rows failed: {self.rows_failed}")

            # Save log
            log_path = self.base_path / f"secondary_techs_update_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            self.save_log(log_path)
            self.log(f"\nLog saved: {log_path}")

            if self.rows_failed > 0:
                self.log("\n⚠ Some rows failed. Please review the log.", "WARNING")
                return 1
            else:
                self.log("\n✓ All changes applied successfully!", "SUCCESS")
                return 0

        except Exception as e:
            self.log(f"\nFATAL ERROR: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return 1


def main():
    try:
        # Paths
        script_dir = Path(__file__).parent
        editor_path = script_dir / "Secondary_Techs_Editor.xlsx"
        base_path = script_dir / "A1_Outputs"
        olade_file_path = script_dir / "Capacidad instalada por fuente - Anual - OLADE.xlsx"

        # Create updater and run
        updater = SecondaryTechsUpdater(editor_path, base_path, olade_file_path)
        return updater.run()

    except Exception as e:
        print(f"\nERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
