import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Define folder paths
INPUT_FOLDER = "OG_csvs_inputs"
OUTPUT_FOLDER = "A1_Outputs"
INPUT_EXCEL_PATH = os.path.join("Miscellaneous", "A-O_Demand.xlsx")
OUTPUT_EXCEL_PATH = os.path.join(OUTPUT_FOLDER, "A-O_Demand.xlsx")

# ISO-3 country code to country name mapping for Latin America and the Caribbean
iso_country_map = {
    "CRI": "Costa Rica", "ARG": "Argentina", "BRA": "Brazil", "COL": "Colombia",
    "PER": "Peru", "CHL": "Chile", "MEX": "Mexico", "VEN": "Venezuela",
    "CUB": "Cuba", "DOM": "Dominican Republic", "PAN": "Panama", "GTM": "Guatemala",
    "ECU": "Ecuador", "BOL": "Bolivia", "URY": "Uruguay", "PRY": "Paraguay",
    "HND": "Honduras", "NIC": "Nicaragua", "SLV": "El Salvador", "JAM": "Jamaica",
    'INT': 'International Markets'
}

# Mapping from code prefix to energy technology description
code_to_energy = {
    'MIN': 'Mining tradable commodity',
    'RNW': 'Mining non-tradable (renewable) commodity',
    'BIO': 'Biomass',
    'GAS': 'Natural Gas',
    'COA': 'Coal',
    'GEO': 'Geothermal',
    'HYD': 'Hydroelectric',
    'OIL': 'Oil',
    'OTH': 'Other',
    'PET': 'Petroleum',
    'SPV': 'Solar Photovoltaic',
    'URN': 'Nuclear',
    'WAV': 'Wave',
    'WAS': 'Waste',
    'WOF': 'Offshore Wind',
    'WON': 'Onshore Wind',
    'CCG': 'Combined Cycle Natural Gas',
    'COG': 'Cogeneration',
    'CSP': 'Concentrated Solar Power',
    'OCG': 'Open Cycle Natural Gas',
    'TRN': 'Transmission technology',
    'LDS': 'Long duration storage',
    'SDS': 'Short duration storage',
    'PWR': 'Power generator',
    'ELC': 'Electricity',
    'BCK': 'Backstop'
}

#-------------------------------------Formated functions--------------------------------------------#
def read_csv_files(input_dir):
    """Reads all CSV files in the given directory and returns a dictionary of DataFrames."""
    data_dict = {}
    for filename in os.listdir(input_dir):
        if filename.endswith(".csv"):
            file_path = os.path.join(input_dir, filename)
            df = pd.read_csv(file_path)
            key = os.path.splitext(filename)[0]
            data_dict[key] = df
    return data_dict

def write_sheet(sheet_name, records, all_years, output_excel_path):
    if not records:
        print(f"[Info] No data to write to sheet '{sheet_name}' in Parametrization file. Skipping.")
        return

    df_out = pd.DataFrame(records)

    for y in all_years:
        if y not in df_out.columns:
            df_out[y] = np.nan

    df_out = df_out.sort_values(by=["Tech.ID", "Parameter.ID"])
    df_out = df_out[[
        "Tech.ID", "Tech", "Tech.Name", "Parameter.ID", "Parameter", "Unit",
        "Projection.Mode", "Projection.Parameter"
    ] + all_years]

    wb = load_workbook(output_excel_path)
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    wb.save(output_excel_path)
    print(f"[Success] Sheet '{sheet_name}' in Parametrization file updated.")

def parse_tech_name(tech):
    iso = tech[6:9]
    region = tech[9:11]
    country = iso_country_map.get(iso, f"Unknown country ({iso})")
    main_code = tech[0:3]
    sub_code = tech[3:6]

    main_desc = code_to_energy.get(main_code, "General technology")
    sub_desc = code_to_energy.get(sub_code, "specific technology")

    # Combinar descripción
    if main_desc != sub_desc:
        base = f"{sub_desc} ({main_desc})"
    else:
        base = main_desc

    name = f"{base} in {country}"

    # Exception to mining
    if not tech.startswith("MIN") and region != "XX":
        name += f", region {region}"

    return name

#--------------------------------------------------------------------------------------------------#

#-------------------------------------Updated functions--------------------------------------------#
def update_demand_profiles(df, output_excel_path, input_excel_path):
    """Updates the Profiles sheet in the given Excel file using the specified DataFrame."""
    # Identify unique years
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    records = []
    for (timeslice, fuel), group in df.groupby(["TIMESLICE", "FUEL"]):
        record = {
            "Timeslices": timeslice,
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel,
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        # Generate descriptive Name field
        iso = fuel[3:6]
        region = fuel[6:8]
        demand = fuel[8:10]
        country = iso_country_map.get(iso, f"Unknown country ({iso})")

        if demand == "01":
            name = f"Output demand of power plants in {country}"
        elif demand == "02":
            name = f"Output demand of transmission lines in {country}"
        else:
            name = f"Unknown demand type for {fuel} in {country}"

        if region != "XX":
            name += f", in region {region}."

        record["Name"] = name

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    # Create DataFrame
    df_timeslices = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Demand/Share", "Fuel/Tech", "Name",
        "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY", "Projection.Mode", "Projection.Parameter"
    ]
    df_timeslices = df_timeslices[fixed_cols + year_cols]

    # Update the Excel sheet
    wb = load_workbook(input_excel_path)
    ws = wb["Profiles"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_timeslices, index=False, header=True):
        ws.append(r)

    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print("[Success] Sheet 'Profiles' in Demand file updated.")

def update_demand_demand_projection(df, output_excel_path, input_excel_path):
    """Updates the Demand_Projection sheet in the given Excel file using the specified DataFrame."""
    # Identify unique years
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    records = []
    for fuel, group in df.groupby("FUEL"):
        record = {
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel,
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        # Generate descriptive Name field
        iso = fuel[3:6]
        region = fuel[6:8]
        demand = fuel[8:10]
        country = iso_country_map.get(iso, f"Unknown country ({iso})")

        if demand == "01":
            name = f"Output demand of power plants in {country}"
        elif demand == "02":
            name = f"Output demand of transmission lines in {country}"
        else:
            name = f"Unknown demand type for {fuel} in {country}"

        if region != "XX":
            name += f", in region {region}."

        record["Name"] = name

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    # Create DataFrame
    df_demand_projection  = pd.DataFrame(records)
    fixed_cols = [
        "Demand/Share", "Fuel/Tech", "Name",
        "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY", "Projection.Mode", "Projection.Parameter"
    ]
    df_demand_projection = df_demand_projection[fixed_cols + year_cols]

    # Update the Excel sheet
    wb = load_workbook(output_excel_path)
    ws = wb["Demand_Projection"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_demand_projection, index=False, header=True):
        ws.append(r)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Demand_Projection' in Demand file updated.")

def update_parametrization_timeslices(df, output_excel_path):
    """Updates Timeslices sheet in A-O_Parametrization.xlsx using CapacityFactor data."""
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    tech_id_map = {}
    tech_id_counter = 1
    records = []

    for (timeslice, tech), group in df.groupby(["TIMESLICE", "TECHNOLOGY"]):
        if tech not in tech_id_map:
            tech_id_map[tech] = tech_id_counter
            tech_id_counter += 1

        record = {
            "Timeslices": timeslice,
            "Tech.ID": tech_id_map[tech],
            "Tech": tech,
            "Parameter.ID": 9,
            "Parameter": "CapacityFactor",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        record["Tech.Name"] = parse_tech_name(tech)

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    df_cap = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Tech.ID", "Tech", "Tech.Name", "Parameter.ID",
        "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"
    ]
    df_cap = df_cap[fixed_cols + year_cols]

    # Solo usa el archivo ubicado en OUTPUT_FOLDER
    wb = load_workbook(output_excel_path)
    ws = wb["Timeslices"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(df_cap, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Timeslices' in Parametrization file updated.")

def update_parametrization_yearsplit(df, output_excel_path):
    """Updates Yearsplit sheet in A-O_Parametrization.xlsx using CapacityFactor data."""
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    # tech_id_map = {}
    # tech_id_counter = 1
    records = []

    for timeslice, group in df.groupby("TIMESLICE"):
        # if tech not in tech_id_map:
        #     tech_id_map[tech] = tech_id_counter
        #     tech_id_counter += 1

        record = {
            "Timeslices": timeslice,
            "Parameter.ID": 14,
            "Parameter": "Yearsplit",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    df_cap = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Parameter.ID",
        "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"
    ]
    df_cap = df_cap[fixed_cols + year_cols]

    # Solo usa el archivo ubicado en OUTPUT_FOLDER
    wb = load_workbook(output_excel_path)
    ws = wb["Yearsplit"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(df_cap, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Yearsplit' in Parametrization file updated.")

def update_parametrization_fixed_horizon_parameters(df_ctau, df_oplife, output_excel_path, input_excel_path):
    """
    Updates the 'Fixed Horizon Parameters' sheet using CapacityToActivityUnit and OperationalLife data.
    Ensures all technologies have entries for all expected parameters, filling with default value = 1.
    Results are sorted by Tech (alphabetical) and then by Parameter.ID (numeric).
    """

    PARAMETERS = [
        ("CapacityToActivityUnit", 1, df_ctau),
        ("OperationalLife", 2, df_oplife)
    ]

    all_techs = set()
    param_data = {}

    for param_name, param_id, df in PARAMETERS:
        param_data[param_name] = {}
        for _, row in df.iterrows():
            tech = row["TECHNOLOGY"]
            value = row["VALUE"]
            param_data[param_name][tech] = value
            all_techs.add(tech)

    tech_ids = {tech: idx + 1 for idx, tech in enumerate(sorted(all_techs))}

    output_rows = []
    for tech in all_techs:
        for param_name, param_id, _ in PARAMETERS:
            value = param_data[param_name].get(tech, 1)
            output_rows.append({
                "Tech.Type": "Primary",
                "Tech.ID": tech_ids[tech],
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Parameter.ID": param_id,
                "Parameter": param_name,
                "Unit": None,
                "Value": value
            })

    df_fixed = pd.DataFrame(output_rows)

    # Apply sorting: first by Tech alphabetically, then by Parameter.ID
    df_fixed = df_fixed.sort_values(by=["Tech", "Parameter.ID"], ascending=[True, True])

    # Write to Excel
    wb = load_workbook(input_excel_path)
    ws = wb["Fixed Horizon Parameters"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_fixed, index=False, header=True):
        ws.append(r)

    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print("[Success] Sheet 'Fixed Horizon Parameters' in Parametrization file updated (sorted).")

def update_parametrization_primary_demand_techs(og_data, output_excel_path):
    PARAMETERS = [
        "CapitalCost", "FixedCost", "VariableCost", "ResidualCapacity",
        "TotalAnnualMaxCapacity", "TotalTechnologyAnnualActivityUpperLimit",
        "TotalTechnologyAnnualActivityLowerLimit", "TotalAnnualMinCapacityInvestment",
        "AvailabilityFactor", "ReserveMarginTagFuel",
        "ReserveMarginTagTechnology", "TotalAnnualMaxCapacityInvestment"
    ]

    PARAMETER_IDS = {name: idx + 1 for idx, name in enumerate(PARAMETERS)}
    primary_records, secondary_records, demand_records = [], [], []
    tech_ids = {}
    tech_counter = 1
    all_years = set()
    techs_by_param = {}

    for param in PARAMETERS:
        if param not in og_data:
            continue
        df = og_data[param]
        key_col = "FUEL" if param == "ReserveMarginTagFuel" else "TECHNOLOGY"
        techs_by_param[param] = set(df[key_col].unique())
        if param != "ReserveMarginTagFuel":
            all_years.update(df["YEAR"].unique())

    all_techs = set().union(*techs_by_param.values())

    for tech in all_techs:
        # is_demand_tech = tech.endswith("02")
        is_demand_tech = tech.startswith("PWRTRN")
        main_prefix = tech[0:3]

        if tech not in tech_ids:
            tech_ids[tech] = tech_counter
            tech_counter += 1

        for param in PARAMETERS:
            if param not in og_data:
                continue

            df = og_data[param]
            key_col = "FUEL" if param == "ReserveMarginTagFuel" else "TECHNOLOGY"
            group = df[df[key_col] == tech]

            if is_demand_tech and param in ["CapitalCost", "FixedCost", "ResidualCapacity"]:
                target = demand_records
            elif is_demand_tech:
                continue
            elif main_prefix in ["MIN", "RNW"]:
                target = primary_records
            else:
                target = secondary_records

            record = {
                "Tech.ID": tech_ids[tech],
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Parameter.ID": PARAMETER_IDS[param],
                "Parameter": param,
                "Unit": None,
                "Projection.Parameter": 0
            }

            if group.empty:
                record["Projection.Mode"] = "EMPTY"
                for y in all_years:
                    record[int(y)] = float("nan")
            else:
                available_years = sorted(group["YEAR"].unique())
                year_values = {int(row["YEAR"]): row["VALUE"] for _, row in group.iterrows()}
                values = [year_values.get(y, float("nan")) for y in available_years]

                non_nan_count = sum(pd.notna(values))
                if non_nan_count == 0:
                    mode = "EMPTY"
                elif non_nan_count == 1 and not pd.isna(values[0]):
                    mode = "Flat"
                elif non_nan_count == len(values):
                    mode = "User defined"
                else:
                    mode = "interpolation"

                record["Projection.Mode"] = mode
                for y in available_years:
                    record[int(y)] = year_values.get(y, float("nan"))

            target.append(record)

    all_years = sorted(all_years)
    write_sheet("Primary Techs", primary_records, all_years, output_excel_path)
    write_sheet("Secondary Techs", secondary_records, all_years, output_excel_path)
    write_sheet("Demand Techs", demand_records, all_years, output_excel_path)

def update_xtra_emissions(og_data, input_excel_path, output_excel_path):
    """
    Updates the 'GHGs' sheet in A-Xtra_Emissions.xlsx using EmissionActivityRatio data.
    Keeps only one row per (TECHNOLOGY, EMISSION), taking the first available value.
    """

    if "EmissionActivityRatio" not in og_data:
        print("[Warning] 'EmissionActivityRatio' not found in OG_Input_Data.")
        return

    df = og_data["EmissionActivityRatio"]

    # Group by (TECHNOLOGY, EMISSION) and keep first value
    grouped = (
        df.groupby(["TECHNOLOGY", "EMISSION"], as_index=False)
        .first()[["TECHNOLOGY", "EMISSION", "VALUE"]]
        .rename(columns={
            "TECHNOLOGY": "Tech",
            "EMISSION": "Emission",
            "VALUE": "EmissionActivityRatio"
        })
    )

    # Add constant unit
    grouped["Unit"] = "MtCO2eq/PJ"
    grouped = grouped[["Tech", "Emission", "EmissionActivityRatio", "Unit"]]

    # Ensure output folder exists
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

    # Load base Excel from input and write into output path
    wb = load_workbook(input_excel_path)
    ws = wb["GHGs"]
    ws.delete_rows(1, ws.max_row)

    from openpyxl.utils.dataframe import dataframe_to_rows
    for row in dataframe_to_rows(grouped, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'GHGs' in Extra Emissions file updated.")





#--------------------------------------------------------------------------------------------------#



def main():
    """Main execution function."""
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    global OG_Input_Data
    OG_Input_Data = read_csv_files(INPUT_FOLDER)

    try:
        update_demand_profiles(
            df=OG_Input_Data["SpecifiedDemandProfile"],
            output_excel_path=OUTPUT_EXCEL_PATH,
            input_excel_path=INPUT_EXCEL_PATH
        )
    except KeyError:
        print("SpecifiedDemandProfile not found in OG_Input_Data.")

    try:
        update_demand_demand_projection(
            df=OG_Input_Data["SpecifiedAnnualDemand"],
            output_excel_path=OUTPUT_EXCEL_PATH,
            input_excel_path=INPUT_EXCEL_PATH
        )
    except KeyError:
        print("SpecifiedAnnualDemand not found in OG_Input_Data.")

    try:
        update_parametrization_fixed_horizon_parameters(
            df_ctau=OG_Input_Data["CapacityToActivityUnit"],
            df_oplife=OG_Input_Data["OperationalLife"],
            output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_Parametrization.xlsx"),
            input_excel_path=os.path.join("Miscellaneous", "A-O_Parametrization.xlsx")
        )
    except KeyError:
        print("CapacityToActivityUnit or OperationalLife not found in OG_Input_Data.")    

    try:
        update_parametrization_timeslices(
            df=OG_Input_Data["CapacityFactor"],
            output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_Parametrization.xlsx")
        )
    except KeyError:
        print("CapacityFactor not found in OG_Input_Data.")
    
    # try:
    update_parametrization_primary_demand_techs(
        og_data=OG_Input_Data,
        output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_Parametrization.xlsx")
    )
    # except KeyError:
    #     print("Parameter not found in OG_Input_Data.")

    try:
        update_parametrization_yearsplit(
            df=OG_Input_Data["YearSplit"],
            output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_Parametrization.xlsx")
        )
    except KeyError:
        print("YearSplit not found in OG_Input_Data.")
        
    try:
        update_xtra_emissions(
            og_data=OG_Input_Data,
            input_excel_path=os.path.join("Miscellaneous", "A-Xtra_Emissions.xlsx"),
            output_excel_path=os.path.join("A2_Extra_Inputs", "A-Xtra_Emissions.xlsx")
        )
    except Exception as e:
        print(f"Failed to update extra emissions file: {e}")


if __name__ == "__main__":
    main()
