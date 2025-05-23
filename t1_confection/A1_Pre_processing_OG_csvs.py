import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Define folder paths
INPUT_FOLDER = "OG_csvs_inputs"
OUTPUT_FOLDER = "A1_Outputs"
INPUT_EXCEL_PATH = os.path.join("Miscellaneous", "A-O_Demand.xlsx")
OUTPUT_EXCEL_PATH = os.path.join(OUTPUT_FOLDER, "A-O_Demand.xlsx")

# ISO-3 country code to country name mapping for Latin America and the Caribbean
iso_country_map = {
    "CRI": "Costa Rica", 
    "ARG": "Argentina", 
    "BRA": "Brazil", 
    "COL": "Colombia",
    "PER": "Peru",
    "CHL": "Chile",
    "MEX": "Mexico",
    "VEN": "Venezuela",
    "CUB": "Cuba",
    "DOM": "Dominican Republic",
    "PAN": "Panama",
    "GTM": "Guatemala",
    "ECU": "Ecuador",
    "BOL": "Bolivia",
    "URY": "Uruguay",
    "PRY": "Paraguay",
    "HND": "Honduras",
    "NIC": "Nicaragua",
    "SLV": "El Salvador",
    "JAM": "Jamaica",
    'INT': 'International Markets'
}

# Mapping from code prefix to energy technology description
code_to_energy = {
    'MIN': 'Mining tradable commodity',
    'RNW': 'Mining non-tradable (renewable) commodity',
    'BIO': 'Biomass',
    'GAS': 'Natural Gas',
    'COA': 'Coal',
    'GEO': 'Geothermal',
    'HYD': 'Hydroelectric',
    'OIL': 'Oil',
    'OTH': 'Other',
    'PET': 'Petroleum',
    'SPV': 'Solar Photovoltaic',
    'URN': 'Nuclear',
    'WAV': 'Wave',
    'WAS': 'Waste',
    'WOF': 'Offshore Wind',
    'WON': 'Onshore Wind',
    'CCG': 'Combined Cycle Natural Gas',
    'COG': 'Cogeneration',
    'CSP': 'Concentrated Solar Power',
    'OCG': 'Open Cycle Natural Gas',
    'TRN': 'Transmission technology',
    'LDS': 'Long duration storage',
    'SDS': 'Short duration storage',
    'PWR': 'Power generator',
    'ELC': 'Electricity',
    'BCK': 'Backstop',
    'CCS': 'Carbon Capture Storage with Coal'
}

#-------------------------------------Formated functions--------------------------------------------#
def read_csv_files(input_dir):
    """Reads all CSV files in the given directory and returns a dictionary of DataFrames."""
    data_dict = {}
    for filename in os.listdir(input_dir):
        if filename.endswith(".csv"):
            file_path = os.path.join(input_dir, filename)
            df = pd.read_csv(file_path)
            key = os.path.splitext(filename)[0]
            data_dict[key] = df
    return data_dict

def write_sheet(sheet_name, records, all_years, output_excel_path):
    if not records:
        print(f"[Info] No data to write to sheet '{sheet_name}' in Parametrization file. Skipping.")
        return

    df_out = pd.DataFrame(records)

    for y in all_years:
        if y not in df_out.columns:
            df_out[y] = np.nan

    df_out = df_out.sort_values(by=["Tech.ID", "Parameter.ID"])
    df_out = df_out[[
        "Tech.ID", "Tech", "Tech.Name", "Parameter.ID", "Parameter", "Unit",
        "Projection.Mode", "Projection.Parameter"
    ] + all_years]

    wb = load_workbook(output_excel_path)
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_out, index=False, header=True):
        ws.append(r)

    wb.save(output_excel_path)
    print(f"[Success] Sheet '{sheet_name}' in Parametrization file updated.")

def parse_tech_name(tech):
    """
    Returns a descriptive name for a technology code with structure interpretation.
    Adds notes based on investability and handles special cases for transmission links.
    """
    main_code = tech[0:3]

    if main_code == "TRN" and len(tech) >= 13:
        iso1 = tech[3:6]
        region1 = tech[6:8]
        iso2 = tech[8:11]
        region2 = tech[11:13]
        country1 = iso_country_map.get(iso1, f"Unknown ({iso1})")
        country2 = iso_country_map.get(iso2, f"Unknown ({iso2})")
        return f"Transmission interconnection from {country1}, region {region1} to {country2}, region {region2}"

    iso = tech[6:9]
    region = tech[9:11]
    country = iso_country_map.get(iso, f"Unknown ({iso})")
    sub_code = tech[3:6]

    main_desc = code_to_energy.get(main_code, "General technology")
    sub_desc = code_to_energy.get(sub_code, "specific technology")

    if main_desc != sub_desc:
        base = f"{sub_desc} ({main_desc})"
    else:
        base = sub_desc

    name = f"{base} {country}"

    # Add region info except for mining
    if not tech.startswith("MIN") and region != "XX":
        name += f", region {region}"
    elif region == "XX":
        name += f", region XX"

    # Add investment status suffix for PWR
    if tech.startswith("PWR"):
        if tech.endswith("00"):
            name += " (can not be invested)"
        elif tech.endswith("01"):
            name += " (can be invested)"

    return name

def parse_fuel_name(fuel):
    """
    Generates a readable name for a fuel code based on its structure.
    Structure format:
    - First 3 characters: fuel type (e.g., OIL, HYD, PET)
    - Characters 3-5: ISO-3 country code
    - Characters 6-7 (if present): region
    - Ending in '01' or '02' is interpreted as output type
    """
    prefix = fuel[0:3]
    iso = fuel[3:6]
    region = fuel[6:8] if len(fuel) >= 8 else None
    suffix = None

    if fuel.endswith("01"):
        suffix = "power plant output"
    elif fuel.endswith("02"):
        suffix = "transmission line output"

    fuel_type = code_to_energy.get(prefix, "Unknown")
    country = iso_country_map.get(iso, f"Unknown ({iso})")

    name_parts = [fuel_type, country]
    if region and region != "XX":
        name_parts.append(f"region {region}")
    elif region == "XX":
        name_parts.append("region XX")
    if suffix:
        name_parts.append(suffix)

    return ", ".join(name_parts)

def assign_tech_type(tech):
    if tech.startswith("MIN") or tech.startswith("RNW"):
        return "Primary"
    elif tech.startswith("PWRTRN"):
        return "Demand"
    else:
        return "Secondary"

#--------------------------------------------------------------------------------------------------#

#-------------------------------------Updated intermediate functions-------------------------------#
def update_demand_profiles(df, output_excel_path, input_excel_path):
    """Updates the Profiles sheet in the given Excel file using the specified DataFrame."""
    # Identify unique years
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    records = []
    for (timeslice, fuel), group in df.groupby(["TIMESLICE", "FUEL"]):
        record = {
            "Timeslices": timeslice,
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel,
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        # Generate descriptive Name field
        iso = fuel[3:6]
        region = fuel[6:8]
        demand = fuel[8:10]
        country = iso_country_map.get(iso, f"Unknown country ({iso})")

        if demand == "01":
            name = f"Output demand of power plants in {country}"
        elif demand == "02":
            name = f"Output demand of transmission lines in {country}"
        else:
            name = f"Unknown demand type for {fuel} in {country}"

        if region != "XX":
            name += f", in region {region}."

        record["Name"] = name

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    # Create DataFrame
    df_timeslices = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Demand/Share", "Fuel/Tech", "Name",
        "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY", "Projection.Mode", "Projection.Parameter"
    ]
    df_timeslices = df_timeslices[fixed_cols + year_cols]

    # Update the Excel sheet
    wb = load_workbook(input_excel_path)
    ws = wb["Profiles"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_timeslices, index=False, header=True):
        ws.append(r)

    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print("[Success] Sheet 'Profiles' in Demand file updated.")

def update_demand_demand_projection(df, output_excel_path, input_excel_path):
    """Updates the Demand_Projection sheet in the given Excel file using the specified DataFrame."""
    # Identify unique years
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    records = []
    for fuel, group in df.groupby("FUEL"):
        record = {
            "Demand/Share": "Demand",
            "Fuel/Tech": fuel,
            "Ref.Cap.BY": "not needed",
            "Ref.OAR.BY": "not needed",
            "Ref.km.BY": "not needed",
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        # Generate descriptive Name field
        iso = fuel[3:6]
        region = fuel[6:8]
        demand = fuel[8:10]
        country = iso_country_map.get(iso, f"Unknown country ({iso})")

        if demand == "01":
            name = f"Output demand of power plants in {country}"
        elif demand == "02":
            name = f"Output demand of transmission lines in {country}"
        else:
            name = f"Unknown demand type for {fuel} in {country}"

        if region != "XX":
            name += f", in region {region}."

        record["Name"] = name

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    # Create DataFrame
    df_demand_projection  = pd.DataFrame(records)
    fixed_cols = [
        "Demand/Share", "Fuel/Tech", "Name",
        "Ref.Cap.BY", "Ref.OAR.BY", "Ref.km.BY", "Projection.Mode", "Projection.Parameter"
    ]
    df_demand_projection = df_demand_projection[fixed_cols + year_cols]

    # Update the Excel sheet
    wb = load_workbook(output_excel_path)
    ws = wb["Demand_Projection"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_demand_projection, index=False, header=True):
        ws.append(r)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Demand_Projection' in Demand file updated.")

def update_parametrization_capacities(df, output_excel_path):
    """Updates Capacities sheet in A-O_Parametrization.xlsx using CapacityFactor data."""
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    tech_id_map = {}
    tech_id_counter = 1
    records = []

    for (timeslice, tech), group in df.groupby(["TIMESLICE", "TECHNOLOGY"]):
        if tech not in tech_id_map:
            tech_id_map[tech] = tech_id_counter
            tech_id_counter += 1

        record = {
            "Timeslices": timeslice,
            "Tech.ID": tech_id_map[tech],
            "Tech": tech,
            "Parameter.ID": 9,
            "Parameter": "CapacityFactor",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        record["Tech.Name"] = parse_tech_name(tech)

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    df_cap = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Tech.ID", "Tech", "Tech.Name", "Parameter.ID",
        "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"
    ]
    df_cap = df_cap[fixed_cols + year_cols]
    # Apply sorting: first by Tech alphabetically, then by Parameter.ID
    df_cap = df_cap.sort_values(by=["Tech.ID", "Timeslices"], ascending=[True, True])


    wb = load_workbook(output_excel_path)
    ws = wb["Capacities"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(df_cap, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Capacities' in Parametrization file updated.")

def update_parametrization_yearsplit(df, output_excel_path):
    """Updates Yearsplit sheet in A-O_Parametrization.xlsx using CapacityFactor data."""
    unique_years = sorted(df["YEAR"].unique())
    year_cols = [str(y) for y in unique_years]

    records = []

    for timeslice, group in df.groupby("TIMESLICE"):

        record = {
            "Timeslices": timeslice,
            "Parameter.ID": 14,
            "Parameter": "Yearsplit",
            "Unit": None,
            "Projection.Mode": "User defined",
            "Projection.Parameter": 0
        }

        for _, row in group.iterrows():
            record[str(row["YEAR"])] = row["VALUE"]

        records.append(record)

    df_cap = pd.DataFrame(records)
    fixed_cols = [
        "Timeslices", "Parameter.ID",
        "Parameter", "Unit", "Projection.Mode", "Projection.Parameter"
    ]
    df_cap = df_cap[fixed_cols + year_cols]

    wb = load_workbook(output_excel_path)
    ws = wb["Yearsplit"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(df_cap, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'Yearsplit' in Parametrization file updated.")

def update_parametrization_fixed_horizon_parameters(df_ctau, df_oplife, output_excel_path, input_excel_path):
    """
    Updates the 'Fixed Horizon Parameters' sheet using CapacityToActivityUnit and OperationalLife data.
    Applies parameter values, fills missing with default = 1, assigns Tech.Type based on naming rules.
    Sorts results by Tech and Parameter.ID before writing to Excel.
    """

    PARAMETERS = [
        ("CapacityToActivityUnit", 1, df_ctau),
        ("OperationalLife", 2, df_oplife)
    ]

    all_techs = set()
    param_data = {}

    for param_name, param_id, df in PARAMETERS:
        param_data[param_name] = {}
        for _, row in df.iterrows():
            tech = row["TECHNOLOGY"]
            value = row["VALUE"]
            param_data[param_name][tech] = value
            all_techs.add(tech)

    tech_ids = {tech: idx + 1 for idx, tech in enumerate(sorted(all_techs))}

    output_rows = []
    for tech in all_techs:
        for param_name, param_id, _ in PARAMETERS:
            value = param_data[param_name].get(tech, 1)
            output_rows.append({
                "Tech.Type": assign_tech_type(tech),
                "Tech.ID": tech_ids[tech],
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Parameter.ID": param_id,
                "Parameter": param_name,
                "Unit": None,
                "Value": value
            })

    df_fixed = pd.DataFrame(output_rows)
    df_fixed = df_fixed.sort_values(by=["Tech", "Parameter.ID"])

    wb = load_workbook(input_excel_path)
    ws = wb["Fixed Horizon Parameters"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_fixed, index=False, header=True):
        ws.append(r)

    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb.save(output_excel_path)
    print("[Success] Sheet 'Fixed Horizon Parameters' in Parametrization updated.")

def update_parametrization_primary_secondary_demand_techs(og_data, output_excel_path):
    """
    Updates Primary, Secondary, and Demand Tech sheets using parameter data.
    Tech type is determined by prefix. Naming logic is conditional:
    - parse_tech_name for MIN, RNW, PWR, TRN
    - parse_fuel_name otherwise
    """
    PARAMETERS = [
        "CapitalCost", "FixedCost", "VariableCost", "ResidualCapacity",
        "TotalAnnualMaxCapacity", "TotalTechnologyAnnualActivityUpperLimit",
        "TotalTechnologyAnnualActivityLowerLimit", "TotalAnnualMinCapacityInvestment",
        "AvailabilityFactor", "ReserveMarginTagFuel",
        "ReserveMarginTagTechnology", "TotalAnnualMaxCapacityInvestment"
    ]

    PARAMETER_IDS = {name: idx + 1 for idx, name in enumerate(PARAMETERS)}
    primary_records, secondary_records, demand_records = [], [], []
    tech_ids = {}
    tech_counter = 1
    all_years = set()
    techs_by_param = {}

    for param in PARAMETERS:
        if param not in og_data:
            continue
        df = og_data[param]
        key_col = "FUEL" if param == "ReserveMarginTagFuel" else "TECHNOLOGY"
        techs_by_param[param] = set(df[key_col].unique())
        if param != "ReserveMarginTagFuel":
            all_years.update(df["YEAR"].unique())

    all_techs = set().union(*techs_by_param.values())

    for tech in all_techs:
        is_demand_tech = tech.startswith("PWRTRN")
        main_prefix = tech[0:3]

        if tech not in tech_ids:
            tech_ids[tech] = tech_counter
            tech_counter += 1

        # Select naming function based on tech prefix
        if main_prefix in ["MIN", "RNW", "PWR", "TRN"]:
            tech_name = parse_tech_name(tech)
        else:
            tech_name = parse_fuel_name(tech)

        for param in PARAMETERS:
            if param not in og_data:
                continue

            df = og_data[param]
            key_col = "FUEL" if param == "ReserveMarginTagFuel" else "TECHNOLOGY"
            group = df[df[key_col] == tech]

            if is_demand_tech and param in ["CapitalCost", "FixedCost", "ResidualCapacity"]:
                target = demand_records
            elif is_demand_tech:
                continue
            elif main_prefix in ["MIN", "RNW"]:
                target = primary_records
            else:
                target = secondary_records

            record = {
                "Tech.ID": tech_ids[tech],
                "Tech": tech,
                "Tech.Name": tech_name,
                "Parameter.ID": PARAMETER_IDS[param],
                "Parameter": param,
                "Unit": None,
                "Projection.Parameter": 0
            }

            if group.empty:
                record["Projection.Mode"] = "EMPTY"
                for y in all_years:
                    record[int(y)] = float("nan")
            else:
                available_years = sorted(group["YEAR"].unique())
                year_values = {int(row["YEAR"]): row["VALUE"] for _, row in group.iterrows()}
                values = [year_values.get(y, float("nan")) for y in available_years]

                non_nan_count = sum(pd.notna(values))
                if non_nan_count == 0:
                    mode = "EMPTY"
                elif non_nan_count == 1 and not pd.isna(values[0]):
                    mode = "Flat"
                elif non_nan_count == len(values):
                    mode = "User defined"
                else:
                    mode = "interpolation"

                record["Projection.Mode"] = mode
                for y in available_years:
                    record[int(y)] = year_values.get(y, float("nan"))

            target.append(record)

    all_years = sorted(all_years)
    write_sheet("Primary Techs", primary_records, all_years, output_excel_path)
    write_sheet("Secondary Techs", secondary_records, all_years, output_excel_path)
    write_sheet("Demand Techs", demand_records, all_years, output_excel_path)



def update_xtra_emissions(og_data, input_excel_path, output_excel_path):
    """
    Updates the 'GHGs' sheet in A-Xtra_Emissions.xlsx using EmissionActivityRatio data.
    Keeps only one row per (TECHNOLOGY, EMISSION, MODE_OF_OPERATION), taking the first available value.
    Includes the Mode_Of_Operation column in the output.
    """
    if "EmissionActivityRatio" not in og_data:
        print("[Warning] 'EmissionActivityRatio' not found in OG_Input_Data.")
        return

    df = og_data["EmissionActivityRatio"]

    # Group by (TECHNOLOGY, EMISSION, MODE_OF_OPERATION) and take the first row
    grouped = (
        df.groupby(["TECHNOLOGY", "EMISSION", "MODE_OF_OPERATION"], as_index=False)
        .first()[["TECHNOLOGY", "EMISSION", "MODE_OF_OPERATION", "VALUE"]]
        .rename(columns={
            "TECHNOLOGY": "Tech",
            "EMISSION": "Emission",
            "MODE_OF_OPERATION": "Mode_Of_Operation",
            "VALUE": "EmissionActivityRatio"
        })
    )

    # Add constant unit
    grouped["Unit"] = "MT"
    grouped = grouped[["Mode_Of_Operation", "Tech", "Emission", "EmissionActivityRatio", "Unit"]]

    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

    wb = load_workbook(input_excel_path)
    ws = wb["GHGs"]
    ws.delete_rows(1, ws.max_row)

    for row in dataframe_to_rows(grouped, index=False, header=True):
        ws.append(row)

    wb.save(output_excel_path)
    print("[Success] Sheet 'GHGs' in Extra Emissions file updated.")
    print("-------------------------------------------------------------------------\n")

def update_model_base_year_primary(og_data, workbook):
    """
    Updates the 'Primary' sheet in the base year model Excel workbook
    using the 'OutputActivityRatio' parameter data from OG_Input_Data.
    Only technologies starting with 'MIN' or 'RNW' are included.
    """
    if "OutputActivityRatio" not in og_data:
        print("[Warning] 'OutputActivityRatio' not found in OG_Input_Data.")
        return

    df = og_data["OutputActivityRatio"]
    df_filtered = df[df["TECHNOLOGY"].str.startswith(("MIN", "RNW"))]

    # Group by unique combinations to extract representative row
    grouped = df_filtered.groupby(["TECHNOLOGY", "FUEL", "MODE_OF_OPERATION"], as_index=False).first()

    records = []
    for _, row in grouped.iterrows():
        tech = row["TECHNOLOGY"]
        fuel = row["FUEL"]
        mode = int(row["MODE_OF_OPERATION"])

        record = {
            "Mode.Operation": mode,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Fuel.O": fuel,
            "Fuel.O.Name": parse_fuel_name(fuel),
            "Value.Fuel.O": 1,  # Always fixed to int(1)
            "Unit.Fuel.O": None
        }
        records.append(record)

    df_final = pd.DataFrame(records)

    # Clear and write to 'Primary' sheet
    ws = workbook["Primary"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Primary' in Base Year Model file updated.")

def update_model_base_year_secondary(og_data, workbook):
    """
    Updates the 'Secondary' sheet in the base year model Excel workbook
    using both 'InputActivityRatio' and 'OutputActivityRatio' data.
    Only includes technologies not starting with 'MIN' or 'RNW', and excludes fuels ending in '02'.
    Each row combines matching input/output records for the same technology and mode of operation.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both parameters: 'InputActivityRatio', 'OutputActivityRatio'.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    df_input = df_input[
        (~df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW"))) #&
        # (~df_input["FUEL"].str.endswith("02"))
    ]
    df_output = df_output[
        (~df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW"))) &
        (~df_output["FUEL"].str.endswith("02"))
    ]

    input_grouped = df_input.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
    output_grouped = df_output.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()

    merged = pd.merge(
        input_grouped, output_grouped,
        on=["TECHNOLOGY", "MODE_OF_OPERATION"],
        suffixes=("_I", "_O")
    )

    records = []
    for _, row in merged.iterrows():
        tech = row["TECHNOLOGY"]
        mode = int(row["MODE_OF_OPERATION"])
        fuel_i = row["FUEL_I"]
        fuel_o = row["FUEL_O"]

        record = {
            "Mode.Operation": mode,
            "Fuel.I": fuel_i,
            "Fuel.I.Name": parse_fuel_name(fuel_i),
            "Value.Fuel.I": 1,
            "Unit.Fuel.I": None,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Fuel.O": fuel_o,
            "Fuel.O.Name": parse_fuel_name(fuel_o),
            "Value.Fuel.O": 1,
            "Unit.Fuel.O": None
        }
        records.append(record)

    df_final = pd.DataFrame(records)

    ws = workbook["Secondary"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Secondary' in Base Year Model file updated.")

def update_model_base_year_demand_techs(og_data, workbook):
    """
    Updates the 'Demand Techs' sheet in the base year model Excel workbook
    using filtered data from 'InputActivityRatio' and 'OutputActivityRatio'.
    Includes only technologies starting with 'PWRTRN' and excludes any starting with 'MIN' or 'RNW'.
    Requires input fuels ending in '01' and output fuels ending in '02'.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both parameters: 'InputActivityRatio', 'OutputActivityRatio'.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    df_input = df_input[
        df_input["TECHNOLOGY"].str.startswith("PWRTRN") &
        (~df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW"))) &
        df_input["FUEL"].str.endswith("01")
    ]
    df_output = df_output[
        df_output["TECHNOLOGY"].str.startswith("PWRTRN") &
        (~df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW"))) &
        df_output["FUEL"].str.endswith("02")
    ]

    input_grouped = df_input.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()
    output_grouped = df_output.groupby(["TECHNOLOGY", "MODE_OF_OPERATION"], as_index=False).first()

    merged = pd.merge(
        input_grouped, output_grouped,
        on=["TECHNOLOGY", "MODE_OF_OPERATION"],
        suffixes=("_I", "_O")
    )

    records = []
    for _, row in merged.iterrows():
        tech = row["TECHNOLOGY"]
        mode = int(row["MODE_OF_OPERATION"])
        fuel_i = row["FUEL_I"]
        fuel_o = row["FUEL_O"]

        record = {
            "Mode.Operation": mode,
            "Fuel.I": fuel_i,
            "Fuel.I.Name": parse_fuel_name(fuel_i),
            "Value.Fuel.I": 1,
            "Unit.Fuel.I": None,
            "Tech": tech,
            "Tech.Name": parse_tech_name(tech),
            "Fuel.O": fuel_o,
            "Fuel.O.Name": parse_fuel_name(fuel_o),
            "Value.Fuel.O": 1,
            "Unit.Fuel.O": None
        }
        records.append(record)

    df_final = pd.DataFrame(records)

    ws = workbook["Demand Techs"]
    ws.delete_rows(1, ws.max_row)

    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Demand Techs' in Base Year Model file updated.")
    
def update_projection_primary(og_data, workbook):
    """
    Updates the 'Primary' sheet in the projection Excel workbook using InputActivityRatio and OutputActivityRatio.
    Filters for technologies starting with 'MIN' or 'RNW'. Sets direction and adapts year columns based on parameter range.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both required parameters for Primary projection.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    # Filter for technologies that start with MIN or RNW
    df_input = df_input[df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW"))]
    df_output = df_output[df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW"))]

    # Determine the union of all years used
    all_years = sorted(set(df_input["YEAR"]).union(df_output["YEAR"]))

    def build_records(df, direction):
        records = []
        for (tech, mode, fuel), group in df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL"]):
            row = {
                "Mode.Operation": int(mode),
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Fuel": fuel,
                "Fuel.Name": parse_fuel_name(fuel),
                "Direction": direction,
                "Projection.Parameter": 0
            }

            year_values = {int(y): v for y, v in zip(group["YEAR"], group["VALUE"])}
            values = [year_values.get(y, np.nan) for y in all_years]

            # Determine Projection.Mode
            non_nan = [v for v in values if pd.notna(v)]
            if not non_nan:
                row["Projection.Mode"] = "EMPTY"
            elif len(non_nan) == 1 and not pd.isna(non_nan[0]):
                row["Projection.Mode"] = "Flat"
            elif len(non_nan) == len(all_years):
                row["Projection.Mode"] = "User defined"
            else:
                row["Projection.Mode"] = "interpolation"

            for y in all_years:
                row[str(y)] = year_values.get(y, np.nan)

            records.append(row)
        return records

    input_records = build_records(df_input, "Input")
    output_records = build_records(df_output, "Output")
    df_final = pd.DataFrame(input_records + output_records)

    # Reorder columns: fixed ones first, then years
    fixed_cols = [
        "Mode.Operation", "Tech", "Tech.Name", "Fuel", "Fuel.Name",
        "Direction", "Projection.Mode", "Projection.Parameter"
    ]
    df_final = df_final[fixed_cols + [str(y) for y in all_years]]
    df_final = df_final.sort_values(by=["Tech", "Direction"])

    # Write to sheet
    ws = workbook["Primary"]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Primary' in Projections file updated.")
    
def update_projection_secondary(og_data, workbook):
    """
    Updates the 'Secondary' sheet in the projection Excel workbook using InputActivityRatio and OutputActivityRatio.
    Includes only technologies that do not start with 'MIN', 'RNW', or 'PWRTRN'.
    Sets direction as 'Input' or 'Output' and adapts year columns to the parameter data.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both required parameters for Secondary projection.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    df_input = df_input[
        ~df_input["TECHNOLOGY"].str.startswith(("MIN", "RNW", "PWRTRN"))
    ]
    df_output = df_output[
        ~df_output["TECHNOLOGY"].str.startswith(("MIN", "RNW", "PWRTRN"))
    ]

    all_years = sorted(set(df_input["YEAR"]).union(df_output["YEAR"]))

    def build_records(df, direction):
        records = []
        for (tech, mode, fuel), group in df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL"]):
            row = {
                "Mode.Operation": int(mode),
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Fuel": fuel,
                "Fuel.Name": parse_fuel_name(fuel),
                "Direction": direction,
                "Projection.Parameter": 0
            }

            year_values = {int(y): v for y, v in zip(group["YEAR"], group["VALUE"])}
            values = [year_values.get(y, np.nan) for y in all_years]

            non_nan = [v for v in values if pd.notna(v)]
            if not non_nan:
                row["Projection.Mode"] = "EMPTY"
            elif len(non_nan) == 1 and not pd.isna(non_nan[0]):
                row["Projection.Mode"] = "Flat"
            elif len(non_nan) == len(all_years):
                row["Projection.Mode"] = "User defined"
            else:
                row["Projection.Mode"] = "interpolation"

            for y in all_years:
                row[str(y)] = year_values.get(y, np.nan)

            records.append(row)
        return records

    input_records = build_records(df_input, "Input")
    output_records = build_records(df_output, "Output")
    df_final = pd.DataFrame(input_records + output_records)

    fixed_cols = [
        "Mode.Operation", "Tech", "Tech.Name", "Fuel", "Fuel.Name",
        "Direction", "Projection.Mode", "Projection.Parameter"
    ]
    df_final = df_final[fixed_cols + [str(y) for y in all_years]]
    df_final = df_final.sort_values(by=["Tech", "Direction"])

    ws = workbook["Secondary"]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Secondary' in Projections file updated.")

def update_projection_demand_techs(og_data, workbook):
    """
    Updates the 'Demand Techs' sheet in the projection Excel workbook using InputActivityRatio and OutputActivityRatio.
    Includes only technologies that start with 'PWRTRN', with input fuels ending in '01' and output fuels ending in '02'.
    Sorts the final result by 'Tech' and 'Direction', and adapts year columns to the parameter data.
    """
    if "InputActivityRatio" not in og_data or "OutputActivityRatio" not in og_data:
        print("[Warning] Missing one or both required parameters for Demand Techs projection.")
        return

    df_input = og_data["InputActivityRatio"]
    df_output = og_data["OutputActivityRatio"]

    df_input = df_input[
        df_input["TECHNOLOGY"].str.startswith("PWRTRN") &
        df_input["FUEL"].str.endswith("01")
    ]
    df_output = df_output[
        df_output["TECHNOLOGY"].str.startswith("PWRTRN") &
        df_output["FUEL"].str.endswith("02")
    ]

    all_years = sorted(set(df_input["YEAR"]).union(df_output["YEAR"]))

    def build_records(df, direction):
        records = []
        for (tech, mode, fuel), group in df.groupby(["TECHNOLOGY", "MODE_OF_OPERATION", "FUEL"]):
            row = {
                "Mode.Operation": int(mode),
                "Tech": tech,
                "Tech.Name": parse_tech_name(tech),
                "Fuel": fuel,
                "Fuel.Name": parse_fuel_name(fuel),
                "Direction": direction,
                "Projection.Parameter": 0
            }

            year_values = {int(y): v for y, v in zip(group["YEAR"], group["VALUE"])}
            values = [year_values.get(y, np.nan) for y in all_years]

            non_nan = [v for v in values if pd.notna(v)]
            if not non_nan:
                row["Projection.Mode"] = "EMPTY"
            elif len(non_nan) == 1 and not pd.isna(non_nan[0]):
                row["Projection.Mode"] = "Flat"
            elif len(non_nan) == len(all_years):
                row["Projection.Mode"] = "User defined"
            else:
                row["Projection.Mode"] = "interpolation"

            for y in all_years:
                row[str(y)] = year_values.get(y, np.nan)

            records.append(row)
        return records

    input_records = build_records(df_input, "Input")
    output_records = build_records(df_output, "Output")
    df_final = pd.DataFrame(input_records + output_records)

    fixed_cols = [
        "Mode.Operation", "Tech", "Tech.Name", "Fuel", "Fuel.Name",
        "Direction", "Projection.Mode", "Projection.Parameter"
    ]
    df_final = df_final[fixed_cols + [str(y) for y in all_years]]
    df_final = df_final.sort_values(by=["Tech", "Direction"])

    ws = workbook["Demand Techs"]
    ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    print("[Success] Sheet 'Demand Techs' in Projections file updated.")
    
def update_yaml_conversions(og_data, yaml_path):
    """
    Updates Conversionls, Conversionld, Conversionlh values in a YAML file using OG_Input_Data.
    Replaces the lists while preserving inline comments and formatting.
    """
    params = ["Conversionls", "Conversionld", "Conversionlh"]

    # Extract values from OG_Input_Data
    replacements = {}
    for param in params:
        if param in og_data:
            values = og_data[param]["VALUE"].astype(int).tolist()
            replacements[param] = values
        else:
            print(f"[Warning] {param} not found in OG_Input_Data.")

    # Read the YAML file as plain text
    with open(yaml_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Update each conversion line by pattern
    updated_lines = []
    for line in lines:
        matched = False
        for param in params:
            pattern = rf"^({param}:\s*)\[[^\]]*\](\s*#.*)$"
            match = re.match(pattern, line)
            if match:
                prefix, suffix = match.groups()
                new_list = ", ".join(map(str, replacements.get(param, [])))
                line = f"{prefix}[{new_list}]{suffix}\n"
                matched = True
                break
        updated_lines.append(line if not matched else line)

    # Write the updated YAML content back
    with open(yaml_path, "w", encoding="utf-8") as f:
        f.writelines(updated_lines)

    print("[Success] Lists: 'Conversionls', 'Conversionlh' and 'Conversionld'\n in MOMF_T1_A file updated.")

def update_yaml_xtra_scen(og_data, yaml_path):
    """
    Updates the xtra_scen block of a YAML file using mapped keys from OG_Input_Data.
    Replaces only the values (inside [] or '') in the corresponding xtra_scen lines.
    Preserves formatting and comments.
    """

    # Mapping of OG_Input_Data keys to xtra_scen keys
    key_map = {
        "REGION": "Region",
        "MODE_OF_OPERATION": "Mode_of_Operation",
        "SEASON": "Season",
        "DAYTYPE": "DayType",
        "DAILYTIMEBRACKET": "DailyTimeBracket",
        "TIMESLICE": "Timeslices"
    }

    # Parameters that must be strings in the YAML list
    force_str_keys = {"Season", "DayType", "DailyTimeBracket", "Timeslices"}

    replacements = {}
    for og_key, yaml_key in key_map.items():
        if og_key in og_data:
            values = og_data[og_key]["VALUE"].tolist()
            if yaml_key == "Region":
                replacements[yaml_key] = str(values[0]) if values else ""
            else:
                if yaml_key in force_str_keys:
                    replacements[yaml_key] = [f"'{str(v)}'" for v in values]
                else:
                    replacements[yaml_key] = [int(v) if isinstance(v, (int, float)) else str(v) for v in values]

    # Read YAML as plain text lines
    with open(yaml_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    updated_lines = []
    for line in lines:
        updated = False
        for yaml_key, new_values in replacements.items():
            if yaml_key == "Region":
                pattern = rf"^(\s*{yaml_key}:\s*)'(.*?)'(.*)$"
                match = re.match(pattern, line)
                if match:
                    prefix, _, suffix = match.groups()
                    line = f"{prefix}'{new_values}'{suffix}\n"
                    updated = True
                    break
            else:
                pattern = rf"^(\s*{yaml_key}:\s*)\[[^\]]*\](.*)$"
                match = re.match(pattern, line)
                if match:
                    prefix, suffix = match.groups()
                    formatted = ", ".join(map(str, new_values))
                    line = f"{prefix}[{formatted}]{suffix}\n"
                    updated = True
                    break
        updated_lines.append(line)

    with open(yaml_path, "w", encoding="utf-8") as f:
        f.writelines(updated_lines)

    print("[Success] Dict 'xtra_scen' in MOMF_T1_A file updated.")

def update_yaml_years(og_data, yaml_path):
    """
    Updates the base_year, initial_year, and final_year fields in the YAML file
    based on the first and last YEAR value found in the OG_Input_Data dictionary.
    Properly clears and replaces quoted year values.
    """

    if "YEAR" not in og_data or "VALUE" not in og_data["YEAR"].columns:
        print("[Warning] 'YEAR' parameter with column 'VALUE' not found in OG_Input_Data.")
        return

    years = sorted(og_data["YEAR"]["VALUE"].unique())
    if not years:
        print("[Warning] YEAR data is empty.")
        return

    base_year = str(int(years[0]))
    final_year = str(int(years[-1]))

    # Read YAML as plain text
    with open(yaml_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    updated_lines = []
    for line in lines:
        if "base_year:" in line:
            line = re.sub(r"(base_year:\s*)['\"].*?['\"]", rf"\1'{base_year}'", line)
        elif "initial_year:" in line:
            line = re.sub(r"(initial_year:\s*)['\"].*?['\"]", rf"\1'{base_year}'", line)
        elif "final_year:" in line:
            line = re.sub(r"(final_year:\s*)['\"].*?['\"]", rf"\1'{final_year}'", line)
        updated_lines.append(line)

    with open(yaml_path, "w", encoding="utf-8") as f:
        f.writelines(updated_lines)

    print("[Success] YAML years variables in MOMF_T1_A file updated.")
#--------------------------------------------------------------------------------------------------#

#-------------------------------------Updated main functions---------------------------------------#

def update_model_base_year(og_data, input_excel_path, output_excel_path):
    """
    Orchestrates the update process for the base year model Excel file.
    Updates the 'Primary', 'Secondary', and 'Demand Techs' sheets using OG_Input_Data.
    Loads the base workbook from input_excel_path and saves to output_excel_path.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb = load_workbook(input_excel_path)

    # Update Primary sheet
    update_model_base_year_primary(og_data, wb)

    # Update Secondary sheet
    update_model_base_year_secondary(og_data, wb)

    # Update Demand Techs sheet
    update_model_base_year_demand_techs(og_data, wb)

    # Save final workbook
    wb.save(output_excel_path)
    print("[Success] Excel file 'Model Base Year' updated.")
    print("-------------------------------------------------------------------------\n")

def update_demand(og_data, input_excel_path, output_excel_path):
    """
    Orchestrates updates to the demand Excel file.
    Applies profile and demand projection updates using OG_Input_Data.
    Assumes all necessary keys exist in og_data.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

    update_demand_profiles(
        df=og_data["SpecifiedDemandProfile"],
        output_excel_path=output_excel_path,
        input_excel_path=input_excel_path
    )

    update_demand_demand_projection(
        df=og_data["SpecifiedAnnualDemand"],
        output_excel_path=output_excel_path,
        input_excel_path=input_excel_path
    )
    
    print("[Success] Excel file 'Demand' updated.")
    print("-------------------------------------------------------------------------\n")
    
def update_parametrization(og_data, output_excel_path, input_excel_path):
    """
    Executes all update routines for the A-O_Parametrization Excel file.
    Applies fixed horizon parameters, timeslices, tech parameter sheets, and yearsplit.
    Assumes all required keys exist in og_data.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)

    update_parametrization_fixed_horizon_parameters(
        df_ctau=og_data["CapacityToActivityUnit"],
        df_oplife=og_data["OperationalLife"],
        output_excel_path=output_excel_path,
        input_excel_path=input_excel_path
    )

    update_parametrization_capacities(
        df=og_data["CapacityFactor"],
        output_excel_path=output_excel_path
    )

    update_parametrization_primary_secondary_demand_techs(
        og_data=og_data,
        output_excel_path=output_excel_path
    )

    update_parametrization_yearsplit(
        df=og_data["YearSplit"],
        output_excel_path=output_excel_path
    )
    
    print("[Success] Excel file 'Parametrization' updated.")
    print("-------------------------------------------------------------------------\n")
    
def update_projections(og_data, input_excel_path, output_excel_path):
    """
    Coordinates the update of the A-O_AR_Projections Excel file.
    Updates the 'Primary', 'Secondary', and 'Demand Techs' sheets using OG_Input_Data.
    Each update adapts year columns based on actual parameter data ranges and sorts by 'Tech' and 'Direction'.
    """
    os.makedirs(os.path.dirname(output_excel_path), exist_ok=True)
    wb = load_workbook(input_excel_path)

    # Update each sheet with sorting logic inside each function
    update_projection_primary(og_data, wb)
    update_projection_secondary(og_data, wb)
    update_projection_demand_techs(og_data, wb)

    wb.save(output_excel_path)
    print("[Success] Excel file 'Projections' updated.")
    print("-------------------------------------------------------------------------\n")

def update_yaml_structure(og_data, yaml_path):
    """
    Executes YAML updates using OG_Input_Data:
    - Updates Conversionls, Conversionld, Conversionlh
    - Updates xtra_scen block
    - Updates base_year, initial_year, final_year
    """
    update_yaml_conversions(og_data, yaml_path)
    update_yaml_xtra_scen(og_data, yaml_path)
    update_yaml_years(og_data, yaml_path)
    
    print("[Success] Yaml file 'MOMF_T1_A' updated.")
    print("-------------------------------------------------------------------------\n")



#--------------------------------------------------------------------------------------------------#



def main():
    """Main execution function."""
    os.makedirs(INPUT_FOLDER, exist_ok=True)
    global OG_Input_Data
    OG_Input_Data = read_csv_files(INPUT_FOLDER)

    # File A-O_Demand.xlsx
    try:
        update_demand(
            og_data=OG_Input_Data,
            input_excel_path=os.path.join("Miscellaneous", "A-O_Demand.xlsx"),
            output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_Demand.xlsx")
        )
    except KeyError as e:
        print(f"[KeyError] Missing key in OG_Input_Data: {e}")
    except Exception as e:
        print(f"[Error] Failed to update demand file: {e}")

    # File A-O_Parametrization.xlsx
    try:
        update_parametrization(
            og_data=OG_Input_Data,
            input_excel_path=os.path.join("Miscellaneous", "A-O_Parametrization.xlsx"),
            output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_Parametrization.xlsx")
        )
    except KeyError as e:
        print(f"[KeyError] Missing key in OG_Input_Data: {e}")
    except Exception as e:
        print(f"[Error] Failed to update parametrization file: {e}")
    
    # File A-Xtra_Emissions.xlsx
    try:
        update_xtra_emissions(
            og_data=OG_Input_Data,
            input_excel_path=os.path.join("Miscellaneous", "A-Xtra_Emissions.xlsx"),
            output_excel_path=os.path.join("A2_Extra_Inputs", "A-Xtra_Emissions.xlsx")
        )
    except KeyError as e:
        print(f"[KeyError] Missing key in OG_Input_Data: {e}")
    except Exception as e:
        print(f"Failed to update extra emissions file: {e}")

    # File A-O_AR_Model_Base_Year.xlsx
    try:
        update_model_base_year(
            og_data=OG_Input_Data,
            input_excel_path=os.path.join("Miscellaneous", "A-O_AR_Model_Base_Year.xlsx"),
            output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_AR_Model_Base_Year.xlsx")
        )
    except KeyError as e:
        print(f"[KeyError] Missing key in OG_Input_Data: {e}")
    except Exception as e:
        print(f"Failed to update model base year file: {e}")

    try:
        update_projections(
            og_data=OG_Input_Data,
            input_excel_path=os.path.join("Miscellaneous", "A-O_AR_Projections.xlsx"),
            output_excel_path=os.path.join(OUTPUT_FOLDER, "A-O_AR_Projections.xlsx")
        )
    except Exception as e:
        print(f"[Error] Failed to update projections file: {e}")
        
    try:
        update_yaml_structure(
            og_data=OG_Input_Data,
            yaml_path="MOMF_T1_A.yaml"
        )
    except Exception as e:
        print(f"[Error] Failed to update YAML structure: {e}")





if __name__ == "__main__":
    main()
